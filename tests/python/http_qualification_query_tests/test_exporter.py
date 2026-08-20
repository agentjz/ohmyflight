from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from .common import APP_ROOT  # noqa: F401
from http_qualification_query.exporter import ResultExporter
from http_qualification_query.models import QueryRecord, QueryResult


class ExporterTests(unittest.TestCase):
    def test_exports_attributed_module_sheets_and_structured_json(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            exporter = ResultExporter(Path(directory), "sample")
            record = QueryRecord(2, "900002", "样例乙", "粘贴第 2 行")
            result = QueryResult(
                page_name="样例乙",
                technical_rows=[{"#": "1", "技术等级代码": "CAP", "技术等级": "机长"}],
                operation_rows=[{"类型": "航线", "运行资格代码": "R1", "运行资格": "样例资格"}],
                basic_info={"出生日期": "2000-01-01"},
                education_rows=[{"学校": "样例院校", "学历": "本科"}],
                work_rows=[{"部门": "样例部门"}],
                training_record_rows=[{"培训科目": "样例培训", "来源页码": "2"}],
                training_experience_rows=[{"训练科目": "样例训练", "来源页码": "1"}],
            )

            exporter.initialize([record], [])
            exporter.write_success(0, record, result)
            paths = exporter.finalize(1, 1, 0, 0, False, "粘贴输入")

            workbook = load_workbook(paths.excel, data_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["处理报告", "技术资料明细", "基础信息", "技术等级", "运行资格", "培训记录", "训练经历", "汇总"],
                )
                self.assertEqual(
                    [cell.value for cell in workbook["基础信息"][1]],
                    ["员工号", "姓名", "分区", "记录序号", "字段", "值"],
                )
                for sheet_name in ("技术等级", "运行资格", "培训记录", "训练经历"):
                    self.assertEqual([cell.value for cell in workbook[sheet_name][1]][:2], ["员工号", "姓名"])
                    self.assertEqual([cell.value for cell in workbook[sheet_name][2]][:2], ["900002", "样例乙"])
                report_values = [cell.value for cell in workbook["处理报告"][2]]
                self.assertEqual(report_values[6:11], [3, 1, 1, 1, 1])
            finally:
                workbook.close()

            payload = json.loads(paths.json.read_text(encoding="utf-8"))
            self.assertEqual(payload["format"], "flight-personnel-info-v1")
            self.assertEqual(payload["summary"]["success"], 1)
            self.assertEqual(payload["people"][0]["employeeId"], "900002")
            self.assertEqual(payload["people"][0]["data"]["training_record_rows"][0]["来源页码"], "2")


if __name__ == "__main__":
    unittest.main()
