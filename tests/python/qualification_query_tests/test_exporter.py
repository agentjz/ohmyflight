from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from .common import APP_DIR  # noqa: F401
from qualification_query.exporter import ResultExporter
from qualification_query.models import QueryRecord, QueryResult


class QualificationExporterTest(unittest.TestCase):
    def test_writes_each_person_to_report_and_detail_sheets(self):
        record = QueryRecord(2, "000001", "测试甲", "粘贴第1条")
        result = QueryResult(
            page_name="测试甲",
            technical_rows=[{"#": "1", "技术等级代码": "CAP", "技术等级": "机长", "机型": "777"}],
            operation_rows=[{"类型": "区域航线资格", "运行资格代码": "R1", "运行资格": "区域资格", "机型": "777"}],
        )
        with tempfile.TemporaryDirectory() as directory:
            exporter = ResultExporter(Path(directory), "unit-run")
            exporter.initialize([record], [])
            exporter.write_success(0, record, result)
            paths = exporter.finalize(total=1, success=1, failed=0, input_errors=0, interrupted=False)

            workbook = load_workbook(paths.excel, data_only=True)
            self.assertEqual(workbook.sheetnames, ["处理报告", "技术资料明细", "汇总"])
            self.assertEqual(workbook["处理报告"]["B2"].value, "000001")
            self.assertEqual(workbook["处理报告"]["G2"].value, 1)
            self.assertEqual(workbook["处理报告"]["H2"].value, 1)
            self.assertEqual(workbook["技术资料明细"].max_row, 3)
            workbook.close()
            self.assertIn("成功人数: 1", paths.report.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()

