from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .common import APP_DIR  # noqa: F401; ensures the app modules are importable
from flight_stats.exporter import ResultExporter
from flight_stats.input_data import QueryRecord
from flight_stats.portal import TableResult


class FlightStatsExporterTest(unittest.TestCase):
    def test_scope_exports_selected_experience_and_landing_count(self):
        with tempfile.TemporaryDirectory() as directory:
            record = QueryRecord("000001", "测试甲", date(2025, 2, 27), date(2026, 8, 18), "粘贴第1行")
            exporter = ResultExporter(Path(directory), "flight-time-run", scope="flight_time")
            exporter.initialize([record], [])
            exporter.write_success(
                0,
                record,
                TableResult(
                    headers=["员工号", "姓名", "飞行时间", "飞行经历", "左座经历", "起落总数"],
                    values={
                        "员工号": "000001",
                        "姓名": "测试甲",
                        "飞行时间": "1269:08",
                        "飞行经历": "613:59",
                        "左座经历": "0",
                        "起落总数": "20",
                    },
                ),
            )

            original = load_workbook(exporter.paths.original, data_only=True)
            stripped = load_workbook(exporter.paths.stripped, data_only=True)
            original_headers = [cell.value for cell in original["查询结果"][1]]
            stripped_headers = [cell.value for cell in stripped["查询结果"][1]]

            self.assertIn("飞行时间", original_headers)
            self.assertIn("起落总数", original_headers)
            self.assertNotIn("飞行经历", original_headers)
            self.assertNotIn("左座经历", original_headers)
            self.assertEqual(stripped_headers, original_headers)
            original.close()
            stripped.close()

    def test_scope_checkboxes_can_combine_experiences(self):
        with tempfile.TemporaryDirectory() as directory:
            record = QueryRecord("000001", "测试甲", date(2025, 2, 27), date(2026, 8, 18), "粘贴第1行")
            exporter = ResultExporter(
                Path(directory),
                "combined-run",
                scope=["flight_time", "left_seat_experience"],
            )
            exporter.initialize([record], [])
            exporter.write_success(
                0,
                record,
                TableResult(
                    headers=["员工号", "姓名", "飞行时间", "飞行经历", "左座经历", "起落总数"],
                    values={
                        "员工号": "000001",
                        "姓名": "测试甲",
                        "飞行时间": "1269:08",
                        "飞行经历": "613:59",
                        "左座经历": "0",
                        "起落总数": "20",
                    },
                ),
            )
            workbook = load_workbook(exporter.paths.original, data_only=True)
            headers = [cell.value for cell in workbook["查询结果"][1]]
            self.assertIn("飞行时间", headers)
            self.assertIn("左座经历", headers)
            self.assertNotIn("飞行经历", headers)
            workbook.close()

    def test_selected_scope_requires_landing_count(self):
        with tempfile.TemporaryDirectory() as directory:
            record = QueryRecord("000001", "测试甲", date(2025, 2, 27), date(2026, 8, 18), "粘贴第1行")
            exporter = ResultExporter(Path(directory), "missing-landing-run", scope="flight_experience")
            exporter.initialize([record], [])
            with self.assertRaisesRegex(ValueError, "起落总数"):
                exporter.write_success(
                    0,
                    record,
                    TableResult(
                        headers=["员工号", "姓名", "飞行经历"],
                        values={"员工号": "000001", "姓名": "测试甲", "飞行经历": "10:30"},
                    ),
                )

    def test_exports_original_and_strips_only_three_requested_fields(self):
        with tempfile.TemporaryDirectory() as directory:
            record = QueryRecord("000001", "测试甲", date(2025, 2, 27), date(2026, 8, 18), "粘贴第1行")
            exporter = ResultExporter(Path(directory), "test-run")
            exporter.initialize([record], [])
            exporter.write_success(
                0,
                record,
                TableResult(
                    headers=[
                        "员工号", "姓名", "飞行时间", "飞行经历", "夜航经历", "左座经历", "右座经历", "起落总数",
                    ],
                    values={
                        "员工号": "000001", "姓名": "测试甲", "飞行时间": "1269:08", "飞行经历": "613:59",
                        "夜航经历": "255:28", "左座经历": "0", "右座经历": "613:59", "起落总数": "20",
                    },
                ),
            )

            original = load_workbook(exporter.paths.original, data_only=True)
            stripped = load_workbook(exporter.paths.stripped, data_only=True)
            original_sheet = original["查询结果"]
            stripped_sheet = stripped["查询结果"]
            headers = [cell.value for cell in original_sheet[1]]

            def value(sheet, header):
                return sheet.cell(2, headers.index(header) + 1).value

            self.assertEqual(value(original_sheet, "飞行时间"), "1269:08")
            self.assertEqual(value(original_sheet, "飞行经历"), "613:59")
            self.assertEqual(value(original_sheet, "左座经历"), "0")
            self.assertEqual(value(stripped_sheet, "飞行时间"), "1269")
            self.assertEqual(value(stripped_sheet, "飞行经历"), "613")
            self.assertEqual(value(stripped_sheet, "左座经历"), "0")
            self.assertEqual(value(stripped_sheet, "夜航经历"), "255:28")
            self.assertEqual(value(stripped_sheet, "右座经历"), "613:59")
            self.assertNotIn("昼间起落", headers)
            self.assertNotIn("夜间起落", headers)
            original.close()
            stripped.close()


if __name__ == "__main__":
    unittest.main()
