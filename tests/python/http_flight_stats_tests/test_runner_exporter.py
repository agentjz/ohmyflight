from __future__ import annotations

import tempfile
import threading
import time
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .common import APP_ROOT  # noqa: F401
from http_flight_stats.exporter import ResultExporter, select_result_headers
from http_flight_stats.models import QueryRecord, TableResult
from http_flight_stats.portal_client import PortalError, PortalSessionExpired
from http_flight_stats.runner import BatchRunner


HEADERS = ["员工号", "姓名", "飞行时间", "飞行经历", "左座经历", "起落总数", "航线起落"]


def make_record(index: int) -> QueryRecord:
    return QueryRecord(
        employee_id=f"{100000 + index:06d}",
        name=f"测试{index}",
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        source=f"第{index}行",
    )


class SerialPortal:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.active = 0
        self.max_active = 0
        self.call_order: list[int] = []

    def query(self, record: QueryRecord) -> TableResult:
        with self.lock:
            self.active += 1
            self.max_active = max(self.max_active, self.active)
            self.call_order.append(int(record.employee_id[-1]))
        try:
            time.sleep(0.01 * (7 - int(record.employee_id[-1])))
            values = {
                "员工号": record.employee_id,
                "姓名": record.name,
                "飞行时间": "120:35",
                "飞行经历": "80:20",
                "左座经历": "40:15",
                "起落总数": "36",
                "航线起落": "30",
            }
            return TableResult(headers=HEADERS, values=values)
        finally:
            with self.lock:
                self.active -= 1


class FailingPortal(SerialPortal):
    def query(self, record: QueryRecord) -> TableResult:
        with self.lock:
            self.active += 1
            self.max_active = max(self.max_active, self.active)
            self.call_order.append(int(record.employee_id[-1]))
        try:
            if record.employee_id.endswith("2"):
                raise PortalError("单人查询失败")
            return TableResult(
                headers=HEADERS,
                values={"员工号": record.employee_id, "姓名": record.name, "飞行时间": "120:35", "飞行经历": "80:20", "左座经历": "40:15", "起落总数": "36", "航线起落": "30"},
            )
        finally:
            with self.lock:
                self.active -= 1


class ExpiringPortal(SerialPortal):
    def query(self, record: QueryRecord) -> TableResult:
        with self.lock:
            self.call_order.append(int(record.employee_id[-1]))
        if record.employee_id.endswith("2"):
            raise PortalSessionExpired("登录凭据已失效")
        return TableResult(headers=HEADERS, values={"员工号": record.employee_id, "姓名": record.name})


class RunnerExporterTests(unittest.TestCase):
    def test_scope_presets_always_include_landings(self) -> None:
        self.assertEqual(
            select_result_headers(HEADERS, ["flight_time"]),
            ["员工号", "姓名", "飞行时间", "起落总数"],
        )
        self.assertEqual(select_result_headers(HEADERS, ["all"]), HEADERS)

    def test_queries_are_strictly_serial_and_keep_input_order(self) -> None:
        records = [make_record(index) for index in range(1, 7)]
        portal = SerialPortal()
        completed_indexes: list[int] = []
        runner = BatchRunner(
            portal=portal,
            stop_event=threading.Event(),
            emit=lambda event: completed_indexes.append(int(event["index"]))
            if event.get("type") == "record_result"
            else None,
        )

        result = runner.run(records, ["all"])

        self.assertEqual(portal.max_active, 1)
        self.assertEqual(portal.call_order, [1, 2, 3, 4, 5, 6])
        self.assertEqual(completed_indexes, list(range(6)))
        self.assertEqual([item.index for item in result.outcomes], list(range(6)))
        self.assertTrue(all(item.status == "成功" for item in result.outcomes))

    def test_exporter_writes_both_files_only_when_batch_finishes(self) -> None:
        records = [make_record(1), make_record(2)]
        runner = BatchRunner(
            portal=SerialPortal(),
            stop_event=threading.Event(),
            emit=lambda _event: None,
        )
        result = runner.run(records, ["flight_time", "left_seat_experience"])

        with tempfile.TemporaryDirectory() as directory:
            exporter = ResultExporter(Path(directory), "test-run", ["flight_time", "left_seat_experience"])
            self.assertFalse(exporter.paths.original.exists())
            self.assertFalse(exporter.paths.stripped.exists())

            exporter.write(records, [], result.outcomes)

            self.assertTrue(exporter.paths.original.is_file())
            self.assertTrue(exporter.paths.stripped.is_file())
            original = load_workbook(exporter.paths.original, data_only=True)
            stripped = load_workbook(exporter.paths.stripped, data_only=True)
            try:
                headers = [cell.value for cell in original["查询结果"][1]]
                self.assertIn("飞行时间", headers)
                self.assertIn("左座经历", headers)
                self.assertIn("起落总数", headers)
                flight_time_column = headers.index("飞行时间") + 1
                left_seat_column = headers.index("左座经历") + 1
                self.assertEqual(original["查询结果"].cell(2, flight_time_column).value, "120:35")
                self.assertEqual(stripped["查询结果"].cell(2, flight_time_column).value, "120")
                self.assertEqual(stripped["查询结果"].cell(2, left_seat_column).value, "40")
            finally:
                original.close()
                stripped.close()

    def test_single_failure_does_not_skip_following_records(self) -> None:
        records = [make_record(index) for index in range(1, 4)]
        portal = FailingPortal()

        result = BatchRunner(portal=portal).run(records, ["all"])

        self.assertEqual(portal.call_order, [1, 2, 3])
        self.assertEqual([item.status for item in result.outcomes], ["成功", "失败", "成功"])

    def test_session_expiry_stops_following_records(self) -> None:
        records = [make_record(index) for index in range(1, 4)]
        portal = ExpiringPortal()

        result = BatchRunner(portal=portal).run(records, ["all"])

        self.assertEqual(portal.call_order, [1, 2])
        self.assertTrue(result.session_expired)
        self.assertEqual([item.status for item in result.outcomes], ["成功", "失败", "失败"])
        self.assertIn("登录凭据已失效", result.outcomes[2].error)


if __name__ == "__main__":
    unittest.main()
