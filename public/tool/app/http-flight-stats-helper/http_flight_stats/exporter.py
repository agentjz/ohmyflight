from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook

from .models import QueryOutcome, QueryRecord


INPUT_HEADERS = ["输入员工号", "输入姓名", "输入开始日期", "输入结束日期"]
STATUS_HEADERS = ["查询状态", "错误说明"]
STRIPPED_TIME_FIELDS = {"飞行时间", "飞行经历", "左座经历"}
SCOPE_FIELDS = {
    "flight_time": "飞行时间",
    "flight_experience": "飞行经历",
    "left_seat_experience": "左座经历",
}
SCOPE_LABELS = {
    "flight_time": "飞行时间+起落数",
    "flight_experience": "飞行经历+起落数",
    "left_seat_experience": "左座经历+起落数",
    "all": "全部数据",
}
ScopeValue = str | Sequence[str]


def normalize_scope(scope: ScopeValue) -> tuple[str, ...]:
    values = [scope] if isinstance(scope, str) else list(scope)
    if not values:
        raise ValueError("至少选择一个查询范围")
    if "all" in values:
        return ("all",)
    normalized = tuple(dict.fromkeys(values))
    if any(value not in SCOPE_FIELDS for value in normalized):
        raise ValueError("不支持的查询范围")
    return normalized


def describe_scope(scope: ScopeValue) -> str:
    normalized = normalize_scope(scope)
    if normalized == ("all",):
        return SCOPE_LABELS["all"]
    return "、".join(SCOPE_LABELS[value] for value in normalized)


def select_result_headers(headers: list[str], scope: ScopeValue) -> list[str]:
    normalized = normalize_scope(scope)
    if normalized == ("all",):
        return list(headers)
    required = [SCOPE_FIELDS[value] for value in normalized] + ["起落总数"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"页面缺少查询字段：{'、'.join(missing)}")
    allowed = {"员工号", "姓名", *required}
    return [header for header in headers if header in allowed]


def strip_minutes(value: object) -> object:
    if not isinstance(value, str):
        return value
    match = re.fullmatch(r"\s*(\d+)\s*[:：]\s*\d{1,2}\s*", value)
    return match.group(1) if match else value


@dataclass(frozen=True)
class OutputPaths:
    original: Path
    stripped: Path


def _save_atomic(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


class ResultExporter:
    def __init__(self, output_directory: Path, run_id: str, scope: ScopeValue):
        self.scope = normalize_scope(scope)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", run_id).strip("-") or "run"
        self.paths = OutputPaths(
            output_directory / f"HTTP飞行经历查询_{safe_name}_原版.xlsx",
            output_directory / f"HTTP飞行经历查询_{safe_name}_去分钟版.xlsx",
        )

    def write(
        self,
        records: list[QueryRecord],
        input_errors: list[str],
        outcomes: list[QueryOutcome],
    ) -> OutputPaths:
        outcome_by_index = {item.index: item for item in outcomes}
        result_headers: list[str] = []
        for item in outcomes:
            if item.status != "成功" or item.result is None:
                continue
            for header in select_result_headers(item.result.headers, self.scope):
                if header not in result_headers:
                    result_headers.append(header)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "查询结果"
        worksheet.append(INPUT_HEADERS + STATUS_HEADERS + result_headers)
        for index, record in enumerate(records):
            outcome = outcome_by_index.get(index)
            status = outcome.status if outcome else "未查询"
            error = outcome.error if outcome else "没有收到查询结果"
            values = outcome.result.values if outcome and outcome.result else {}
            worksheet.append(
                [
                    record.employee_id,
                    record.name,
                    record.start_date,
                    record.end_date,
                    status,
                    error,
                    *[values.get(header, "") for header in result_headers],
                ]
            )
            worksheet.cell(worksheet.max_row, 3).number_format = "yyyy-mm-dd"
            worksheet.cell(worksheet.max_row, 4).number_format = "yyyy-mm-dd"
        if input_errors:
            error_sheet = workbook.create_sheet("输入错误")
            error_sheet.append(["来源", "错误说明"])
            for index, message in enumerate(input_errors, start=1):
                error_sheet.append([index, message])
        _save_atomic(workbook, self.paths.original)
        workbook.close()

        stripped = load_workbook(self.paths.original)
        try:
            stripped_sheet = stripped["查询结果"]
            header_map = {str(cell.value or "").strip(): cell.column for cell in stripped_sheet[1]}
            for header in STRIPPED_TIME_FIELDS:
                column = header_map.get(header)
                if not column:
                    continue
                for row_number in range(2, stripped_sheet.max_row + 1):
                    cell = stripped_sheet.cell(row_number, column)
                    cell.value = strip_minutes(cell.value)
            _save_atomic(stripped, self.paths.stripped)
        finally:
            stripped.close()
        return self.paths

