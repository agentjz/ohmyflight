from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import InputPayload, QueryRecord


DATE_PATTERN = re.compile(
    r"(?P<year>\d{4})\s*(?:年|[-/.])\s*(?P<month>\d{1,2})\s*(?:月|[-/.])\s*(?P<day>\d{1,2})\s*(?:日)?"
)
EMPLOYEE_PATTERN = re.compile(r"(?<!\d)(\d{6})(?!\d)")
EMPLOYEE_HEADERS = {"员工号", "工号", "员工编号"}
NAME_HEADERS = {"姓名", "员工姓名"}
START_HEADERS = {"开始日期", "起始时间", "开始时间"}
END_HEADERS = {"结束日期", "截止时间", "结束时间"}


def normalize_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    match = DATE_PATTERN.search(str(value).strip())
    if not match:
        return None
    try:
        return date(int(match.group("year")), int(match.group("month")), int(match.group("day")))
    except ValueError:
        return None


def normalize_employee_id(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0+", text):
            text = text.split(".", 1)[0]
    if not text.isdigit() or len(text) > 6:
        return None
    return text.zfill(6)


def _record_from_values(
    employee_value: Any,
    name_value: Any,
    start_value: Any,
    end_value: Any,
    source: str,
) -> QueryRecord:
    employee_id = normalize_employee_id(employee_value)
    if not employee_id:
        raise ValueError("员工号不是 6 位数字")
    start_date = normalize_date(start_value)
    if not start_date:
        raise ValueError("开始日期格式无效")
    end_date = normalize_date(end_value) if end_value not in (None, "") else start_date
    if not end_date:
        raise ValueError("结束日期格式无效")
    if end_date < start_date:
        raise ValueError("结束日期早于开始日期")
    return QueryRecord(employee_id, str(name_value or "").strip(), start_date, end_date, source)


def _split_pasted_lines(text: str) -> list[str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        return [part.strip() for part in re.split(r"(?=\d{6})", lines[0]) if part.strip()]
    return lines


def parse_pasted_records(text: str) -> tuple[list[QueryRecord], list[str]]:
    records: list[QueryRecord] = []
    errors: list[str] = []
    seen: set[str] = set()
    for index, line in enumerate(_split_pasted_lines(text), start=1):
        employee_match = EMPLOYEE_PATTERN.search(line)
        if not employee_match:
            errors.append(f"第{index}条：未识别员工号")
            continue
        employee_id = employee_match.group(1)
        tail = line[employee_match.end() :].strip()
        name_match = re.match(r"([\u4e00-\u9fff]{2,8})", tail)
        name = name_match.group(1) if name_match else ""
        dates = [normalize_date(match.group(0)) for match in DATE_PATTERN.finditer(line)]
        parsed_dates = [item for item in dates if item]
        if not parsed_dates:
            errors.append(f"第{index}条：未识别日期")
            continue
        try:
            record = _record_from_values(
                employee_id,
                name,
                parsed_dates[0],
                parsed_dates[1] if len(parsed_dates) > 1 else parsed_dates[0],
                f"粘贴第{index}条",
            )
        except ValueError as error:
            errors.append(f"第{index}条：{error}")
            continue
        if record.employee_id in seen:
            errors.append(f"第{index}条：重复员工号，已跳过")
            continue
        seen.add(record.employee_id)
        records.append(record)
    return records, errors


def _header_indexes(headers: list[str]) -> dict[str, int]:
    normalized = [str(value or "").strip() for value in headers]

    def find(candidates: set[str]) -> int | None:
        return next((index for index, header in enumerate(normalized) if header in candidates), None)

    indexes = {
        "employee": find(EMPLOYEE_HEADERS),
        "name": find(NAME_HEADERS),
        "start": find(START_HEADERS),
        "end": find(END_HEADERS),
    }
    missing = [name for name, index in indexes.items() if index is None and name != "name"]
    if missing:
        raise ValueError(f"Excel 缺少必需表头：{', '.join(missing)}")
    return {name: index for name, index in indexes.items() if index is not None}


def read_excel_records(source: str | Path | bytes) -> tuple[list[QueryRecord], list[str]]:
    workbook_source: str | Path | BytesIO = BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(workbook_source, data_only=True)
    records: list[QueryRecord] = []
    errors: list[str] = []
    seen: set[str] = set()
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            return [], ["Excel 没有表头"]
        indexes = _header_indexes([str(value or "").strip() for value in rows[0]])
        for row_number, row in enumerate(rows[1:], start=2):
            if not row or not any(value not in (None, "") for value in row):
                continue

            def get(key: str) -> Any:
                index = indexes[key]
                return row[index] if index < len(row) else None

            try:
                record = _record_from_values(
                    get("employee"),
                    get("name") if "name" in indexes else "",
                    get("start"),
                    get("end"),
                    f"Excel 第{row_number}行",
                )
            except ValueError as error:
                errors.append(f"第{row_number}行：{error}")
                continue
            if record.employee_id in seen:
                errors.append(f"第{row_number}行：重复员工号，已跳过")
                continue
            seen.add(record.employee_id)
            records.append(record)
    finally:
        workbook.close()
    return records, errors


def read_input(payload: InputPayload) -> tuple[list[QueryRecord], list[str]]:
    if payload.input_mode == "excel":
        if not payload.excel_bytes:
            raise ValueError("请选择 Excel 文件")
        return read_excel_records(payload.excel_bytes)
    if payload.input_mode == "paste":
        if not payload.pasted_text.strip():
            raise ValueError("请粘贴查询数据")
        return parse_pasted_records(payload.pasted_text)
    raise ValueError(f"不支持的输入方式：{payload.input_mode}")

