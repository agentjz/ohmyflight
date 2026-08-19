"""Immediate, atomic Excel persistence for HTTP lock-entry evidence."""

from __future__ import annotations

import os
import tempfile
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


RESULT_HEADERS = [
    ("序号", "index"),
    ("片段", "segmentIndex"),
    ("来源行", "sourceRow"),
    ("员工号", "employeeId"),
    ("输入姓名", "expectedName"),
    ("门户姓名", "name"),
    ("输入锁班类型", "inputType"),
    ("实际锁班类型", "actualType"),
    ("输入开始日期", "inputStartDate"),
    ("输入结束日期", "inputEndDate"),
    ("实际开始日期时间", "actualStartDateTime"),
    ("实际结束日期时间", "actualEndDateTime"),
    ("时间模式", "timeMode"),
    ("处理状态", "status"),
    ("门户结果", "portalStatus"),
    ("尝试次数", "attempt"),
    ("冲突回退", "recovery"),
    ("锁班备注", "remark"),
    ("说明", "message"),
    ("记录时间", "recordedAt"),
]


class ResultStore:
    def __init__(self, output_root: str | Path, run_id: str, mode: str):
        self.directory = Path(output_root).resolve() / run_id
        self.directory.mkdir(parents=True, exist_ok=True)
        self.path = str(self.directory / f"HTTP锁班结果_{run_id}.xlsx")
        self.mode = mode
        self._lock = threading.RLock()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "录入结果"
        sheet.append([label for label, _key in RESULT_HEADERS])
        sheet.freeze_panes = "A2"
        self._save_workbook(workbook)

    def _save_workbook(self, workbook: Workbook) -> None:
        file_descriptor, temporary_path = tempfile.mkstemp(
            prefix="http-lock-result-",
            suffix=".xlsx",
            dir=self.directory,
        )
        os.close(file_descriptor)
        try:
            workbook.save(temporary_path)
            workbook.close()
            os.replace(temporary_path, self.path)
        finally:
            if os.path.exists(temporary_path):
                os.unlink(temporary_path)

    def append(self, row: dict[str, object]) -> None:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook["录入结果"]
            saved_row = dict(row)
            saved_row.setdefault("recordedAt", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            sheet.append([saved_row.get(key, "") for _label, key in RESULT_HEADERS])
            self._save_workbook(workbook)
