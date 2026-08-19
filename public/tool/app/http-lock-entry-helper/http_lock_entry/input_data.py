"""Excel and pasted-text normalization for the HTTP lock workbench."""

from __future__ import annotations

import calendar
import re
from datetime import date, datetime, time, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .metadata import normalize_lock_type, normalize_text
from .models import InputPayload, PortalMetadata, WorkbenchMode


REQUIRED_HEADERS = {
    "员工号": ("员工号", "工号"),
    "姓名": ("姓名",),
    "锁班类型": ("锁班类型", "请假类型"),
    "开始日期": ("开始日期",),
    "结束日期": ("结束日期",),
}

OPTIONAL_HEADERS = {
    "时间模式": ("时间模式", "锁班时间类型"),
    "开始时间": ("开始时间",),
    "结束时间": ("结束时间",),
    "月份": ("月份", "锁班月份"),
    "锁班日期": ("锁班日期", "日期列表"),
    "备注": ("备注", "锁班原因"),
}

SMART_LEAVE_TYPES = {"RECU_LVE", "ALV_FD"}


def parse_whitelist(text: str) -> set[str]:
    return set(re.findall(r"\d{6}", str(text or "")))


def normalize_employee_id(value: object) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_business_date(value: object) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        compact = str(int(value)) if float(value).is_integer() else ""
        if re.fullmatch(r"(?:19|20)\d{6}", compact):
            value = compact
        else:
            try:
                return from_excel(value).strftime("%Y-%m-%d")
            except (TypeError, ValueError, OverflowError):
                return ""
    text = str(value).strip()
    compact_match = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", text)
    if compact_match:
        text = "-".join(compact_match.groups())
    separated = re.fullmatch(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if not separated:
        return ""
    year, month, day = (int(part) for part in separated.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def normalize_time(value: object) -> str:
    if value is None or value == "" or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.strftime("%H:%M")
        except (TypeError, ValueError, OverflowError):
            return ""
    match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)", str(value).strip())
    return f"{int(match.group(1)):02d}:{match.group(2)}" if match else ""


def normalize_time_mode(value: object) -> int:
    text = normalize_text(value)
    if not text or text in {"1", "按时间段"}:
        return 1
    if text in {"2", "按月份", "按月份(频率)", "按月份（频率）"}:
        return 2
    return 0


def normalize_month(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    match = re.fullmatch(r"(\d{4})[-/.](\d{1,2})", normalize_text(value))
    if not match:
        return ""
    year, month = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}" if 1 <= month <= 12 else ""


def parse_selected_days(value: object) -> list[int]:
    if value is None or value == "":
        return []
    if isinstance(value, (date, datetime)):
        return [value.day]
    if isinstance(value, (int, float)) and float(value).is_integer() and 1 <= int(value) <= 31:
        return [int(value)]
    days = [int(item) for item in re.findall(r"(?<!\d)(\d{1,2})(?!\d)", str(value))]
    return sorted(set(day for day in days if 1 <= day <= 31))


def _header_indexes(header_row: tuple[object, ...]) -> tuple[dict[str, int], list[str]]:
    headers = [normalize_text(value) for value in header_row]
    indexes: dict[str, int] = {}
    missing = []
    for field, aliases in REQUIRED_HEADERS.items():
        index = next((headers.index(alias) for alias in aliases if alias in headers), None)
        if index is None:
            missing.append(field)
        else:
            indexes[field] = index
    for field, aliases in OPTIONAL_HEADERS.items():
        index = next((headers.index(alias) for alias in aliases if alias in headers), None)
        if index is not None:
            indexes[field] = index
    return indexes, missing


def _value(row: tuple[object, ...], indexes: dict[str, int], field: str) -> object:
    index = indexes.get(field)
    return row[index] if index is not None and index < len(row) else None


def _finalize_record(
    raw: dict[str, object],
    metadata: PortalMetadata,
    mode: WorkbenchMode,
) -> tuple[dict[str, object] | None, str]:
    employee_id = normalize_employee_id(raw.get("员工号"))
    if not re.fullmatch(r"\d{6}", employee_id):
        return None, f"员工号格式错误 [{employee_id}]"

    lock_type = normalize_lock_type(raw.get("锁班类型"), metadata)
    if not lock_type:
        return None, f"未识别当前门户锁班类型 [{normalize_text(raw.get('锁班类型'))}]"
    type_metadata = metadata.lock_types[lock_type]

    time_mode = normalize_time_mode(raw.get("时间模式"))
    if time_mode not in {1, 2}:
        return None, f"锁班时间类型错误 [{normalize_text(raw.get('时间模式'))}]"
    if time_mode == 2 and not type_metadata.supports_monthly:
        return None, f"当前门户类型 [{type_metadata.description}] 只允许按时间段"

    start_date = normalize_business_date(raw.get("开始日期"))
    end_date = normalize_business_date(raw.get("结束日期")) or start_date
    start_time = normalize_time(raw.get("开始时间")) or type_metadata.default_start_time
    end_time = normalize_time(raw.get("结束时间")) or type_metadata.default_end_time
    if not start_time or not end_time:
        return None, "开始时间或结束时间格式错误"

    month = normalize_month(raw.get("月份"))
    selected_days = parse_selected_days(raw.get("锁班日期"))

    if time_mode == 1:
        if not start_date or not end_date:
            return None, "开始日期或结束日期格式错误"
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
        if end < start:
            return None, "结束日期不得早于开始日期"
        if start == end and start_time >= end_time:
            return None, "同一天的开始时间必须早于结束时间"
        if (
            end.day - start.day == 1
            and type_metadata.date_split_flag == "0"
            and (datetime.combine(end, time.fromisoformat(end_time)) - datetime.combine(start, time.fromisoformat(start_time))).total_seconds() <= 8 * 3600
        ):
            return None, "相邻日期的开始时间和结束时间之间必须超过8小时"
        if mode == "smart" and lock_type in SMART_LEAVE_TYPES and start.year != end.year:
            return None, "智能路由暂不处理跨自然年记录"
    else:
        if not month and start_date and end_date:
            start = date.fromisoformat(start_date)
            end = date.fromisoformat(end_date)
            if start.strftime("%Y-%m") == end.strftime("%Y-%m"):
                month = start.strftime("%Y-%m")
                if not selected_days:
                    selected_days = [(start + timedelta(days=offset)).day for offset in range((end - start).days + 1)]
        if not month:
            return None, "按月份模式缺少月份"
        if not selected_days:
            return None, "按月份模式缺少锁班日期"
        year, month_number = (int(part) for part in month.split("-"))
        max_day = calendar.monthrange(year, month_number)[1]
        if any(day > max_day for day in selected_days):
            return None, f"{month}不存在所选日期"
        if start_time >= end_time:
            return None, "按月份模式的开始时间必须早于结束时间"

    remark = normalize_text(raw.get("备注"))
    if len(remark) > 60:
        return None, "锁班原因不能超过60个字符"

    return {
        "员工号": employee_id,
        "姓名": normalize_text(raw.get("姓名")) or None,
        "请假类型": lock_type,
        "开始日期": start_date,
        "结束日期": end_date,
        "时间模式": time_mode,
        "开始时间": start_time,
        "结束时间": end_time,
        "月份": month,
        "日期列表": selected_days,
        "备注": remark,
        "来源行": raw.get("来源行", ""),
    }, ""


def parse_excel(
    content: bytes,
    metadata: PortalMetadata,
    mode: WorkbenchMode,
    whitelist: set[str] | None = None,
) -> tuple[list[dict[str, object]], list[str]]:
    records: list[dict[str, object]] = []
    errors: list[str] = []
    workbook = load_workbook(BytesIO(content), data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        indexes, missing = _header_indexes(header_row)
        if missing:
            return [], [f"Excel缺少表头: {', '.join(missing)}"]
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(value is not None and value != "" for value in row):
                continue
            raw = {
                "员工号": _value(row, indexes, "员工号"),
                "姓名": _value(row, indexes, "姓名"),
                "锁班类型": _value(row, indexes, "锁班类型"),
                "开始日期": _value(row, indexes, "开始日期"),
                "结束日期": _value(row, indexes, "结束日期"),
                "时间模式": _value(row, indexes, "时间模式"),
                "开始时间": _value(row, indexes, "开始时间"),
                "结束时间": _value(row, indexes, "结束时间"),
                "月份": _value(row, indexes, "月份"),
                "锁班日期": _value(row, indexes, "锁班日期"),
                "备注": _value(row, indexes, "备注"),
                "来源行": row_number,
            }
            record, problem = _finalize_record(raw, metadata, mode)
            if problem:
                errors.append(f"第{row_number}行: {problem}")
            elif whitelist and record and record["员工号"] not in whitelist:
                continue
            elif record:
                records.append(record)
    finally:
        workbook.close()
    return records, errors


def _parse_delimited_line(line: str) -> dict[str, object] | None:
    delimiter = "\t" if "\t" in line else "|" if "|" in line else ""
    if not delimiter:
        return None
    parts = [part.strip() for part in line.split(delimiter)]
    if len(parts) < 5:
        return None
    return {
        "员工号": parts[0],
        "姓名": parts[1],
        "锁班类型": parts[2],
        "开始日期": parts[3],
        "结束日期": parts[4],
        "开始时间": parts[5] if len(parts) > 5 else "",
        "结束时间": parts[6] if len(parts) > 6 else "",
        "备注": parts[7] if len(parts) > 7 else "",
    }


def _parse_freeform_line(line: str) -> dict[str, object]:
    employee = re.search(r"(?<!\d)(\d{6})(?!\d)", line)
    name = re.search(r"\d{6}\s*([\u4e00-\u9fa5]{2,4})", line)
    dates = re.findall(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", line)
    times = re.findall(r"(?<!\d)(?:[01]?\d|2[0-3]):[0-5]\d(?!\d)", line)
    remark = re.search(r"(?:备注|锁班原因)\s*[=:：]\s*(.+)$", line)
    return {
        "员工号": employee.group(1) if employee else "",
        "姓名": name.group(1) if name else "",
        "锁班类型": line,
        "开始日期": dates[0] if dates else "",
        "结束日期": dates[1] if len(dates) > 1 else dates[0] if dates else "",
        "开始时间": times[0] if times else "",
        "结束时间": times[1] if len(times) > 1 else "",
        "备注": remark.group(1).strip() if remark else "",
    }


def parse_pasted_text(
    text: str,
    metadata: PortalMetadata,
    mode: WorkbenchMode,
    whitelist: set[str] | None = None,
) -> tuple[list[dict[str, object]], list[str]]:
    records: list[dict[str, object]] = []
    errors: list[str] = []
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    for index, line in enumerate(lines, start=1):
        raw = _parse_delimited_line(line) or _parse_freeform_line(line)
        raw["来源行"] = index
        record, problem = _finalize_record(raw, metadata, mode)
        if problem:
            errors.append(f"第{index}条: {problem}")
        elif whitelist and record and record["员工号"] not in whitelist:
            continue
        elif record:
            records.append(record)
    return records, errors


def read_input(
    payload: InputPayload,
    metadata: PortalMetadata,
    mode: WorkbenchMode,
) -> tuple[list[dict[str, object]], list[str]]:
    whitelist = parse_whitelist(payload.whitelist_text) or None
    if payload.input_mode == "excel":
        if not payload.excel_bytes:
            raise ValueError("请选择 Excel 文件")
        return parse_excel(payload.excel_bytes, metadata, mode, whitelist)
    if payload.input_mode == "paste":
        if not payload.pasted_text.strip():
            raise ValueError("请粘贴锁班数据")
        return parse_pasted_text(payload.pasted_text, metadata, mode, whitelist)
    raise ValueError(f"不支持的输入方式：{payload.input_mode}")
