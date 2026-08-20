from __future__ import annotations

import json
import os
import re
import uuid
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .models import InputIssue, OutputPaths, QueryRecord, QueryResult


DETAIL_HEADERS = (
    "员工号",
    "姓名",
    "资料类型",
    "页面序号",
    "类型",
    "代码",
    "名称",
    "水平等级",
    "机型",
    "生效时间",
    "失效时间",
    "对应检查记录",
    "数据来源",
    "备注",
    "抓取状态",
    "说明",
    "抓取时间",
)
REPORT_HEADERS = (
    "输入行号",
    "员工号",
    "输入姓名",
    "页面姓名",
    "员工号匹配",
    "姓名匹配",
    "基础信息条数",
    "技术等级条数",
    "运行资格条数",
    "培训记录条数",
    "训练经历条数",
    "抓取状态",
    "说明",
    "查询时间",
)
SUMMARY_HEADERS = ("项目", "值")
TECHNICAL_HEADERS = ("#", "技术等级代码", "技术等级", "水平等级", "机型", "生效时间", "失效时间", "对应检查记录", "数据来源")
OPERATION_HEADERS = ("类型", "运行资格代码", "运行资格", "水平等级", "机型", "生效时间", "失效时间", "备注")
TRAINING_RECORD_HEADERS = ("选择", "培训科目", "培训机型", "培训课时", "培训地点", "经办人", "教员", "培训时间", "培训結束时间", "训练结果", "考试成绩", "上传")
TRAINING_EXPERIENCE_HEADERS = ("全选", "序号", "训练日期", "训练机型", "训练科目", "类型", "检查单", "结论", "上传", "审批过程", "证书下载")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_name(value: str) -> str:
    text = re.sub(r"[（(][^）)]*[）)]", "", str(value or ""))
    return re.sub(r"\s+", "", text)


def _safe_run_name(run_id: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", run_id).strip("-") or "run"


def _save_atomic(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


class ResultExporter:
    def __init__(self, output_directory: Path, run_id: str):
        output_directory.mkdir(parents=True, exist_ok=True)
        name = _safe_run_name(run_id)
        self.paths = OutputPaths(
            excel=output_directory / f"飞行人员信息查询_{name}.xlsx",
            report=output_directory / f"飞行人员信息查询_{name}.txt",
            json=output_directory / f"飞行人员信息查询_{name}.json",
        )
        self._json_people: list[dict[str, object]] = []
        self._json_summary: dict[str, object] = {}

    def initialize(self, records: list[QueryRecord], input_issues: list[InputIssue]) -> None:
        workbook = Workbook()
        report_sheet = workbook.active
        report_sheet.title = "处理报告"
        detail_sheet = workbook.create_sheet("技术资料明细")
        basic_sheet = workbook.create_sheet("基础信息")
        technical_sheet = workbook.create_sheet("技术等级")
        operation_sheet = workbook.create_sheet("运行资格")
        training_sheet = workbook.create_sheet("培训记录")
        experience_sheet = workbook.create_sheet("训练经历")
        summary_sheet = workbook.create_sheet("汇总")
        report_sheet.append(list(REPORT_HEADERS))
        detail_sheet.append(list(DETAIL_HEADERS))
        basic_sheet.append(["员工号", "姓名", "分区", "记录序号", "字段", "值"])
        technical_sheet.append(["员工号", "姓名", *TECHNICAL_HEADERS])
        operation_sheet.append(["员工号", "姓名", *OPERATION_HEADERS])
        training_sheet.append(["员工号", "姓名", *TRAINING_RECORD_HEADERS, "来源页码"])
        experience_sheet.append(["员工号", "姓名", *TRAINING_EXPERIENCE_HEADERS, "来源页码"])
        summary_sheet.append(list(SUMMARY_HEADERS))
        for worksheet in (report_sheet, detail_sheet, basic_sheet, technical_sheet, operation_sheet, training_sheet, experience_sheet, summary_sheet):
            _style_header(worksheet)

        for column, width in {
            "A": 11, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
            "G": 16, "H": 16, "I": 14, "J": 48, "K": 20,
        }.items():
            report_sheet.column_dimensions[column].width = width
        for column, width in {
            "A": 12, "B": 12, "C": 12, "D": 10, "E": 18, "F": 16, "G": 34,
            "H": 12, "I": 10, "J": 14, "K": 14, "L": 18, "M": 12, "N": 12,
            "O": 12, "P": 48, "Q": 20,
        }.items():
            detail_sheet.column_dimensions[column].width = width
        summary_sheet.column_dimensions["A"].width = 20
        summary_sheet.column_dimensions["B"].width = 90

        for issue in input_issues:
            report_sheet.append([
                issue.input_row,
                issue.employee_id,
                issue.name,
                "",
                "否",
                "未查询",
                0,
                0,
                0,
                0,
                0,
                "输入错误",
                issue.message,
                now_text(),
            ])
        _save_atomic(workbook, self.paths.excel)
        workbook.close()
        self._write_json()

    def _write_json(self) -> None:
        payload = {
            "format": "flight-personnel-info-v1",
            "summary": self._json_summary,
            "people": self._json_people,
        }
        temporary = self.paths.json.with_suffix(".tmp.json")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temporary, self.paths.json)

    @staticmethod
    def _detail_rows(record: QueryRecord, result: QueryResult) -> list[list[str]]:
        name = result.page_name or record.name
        captured_at = now_text()
        rows: list[list[str]] = []
        for row in result.technical_rows:
            rows.append([
                record.employee_id, name, "技术等级", row.get("#", ""), "", row.get("技术等级代码", ""),
                row.get("技术等级", ""), row.get("水平等级", ""), row.get("机型", ""),
                row.get("生效时间", ""), row.get("失效时间", ""), row.get("对应检查记录", ""),
                row.get("数据来源", ""), "", "成功", "", captured_at,
            ])
        for index, row in enumerate(result.operation_rows, start=1):
            rows.append([
                record.employee_id, name, "运行资格", str(index), row.get("类型", ""), row.get("运行资格代码", ""),
                row.get("运行资格", ""), row.get("水平等级", ""), row.get("机型", ""),
                row.get("生效时间", ""), row.get("失效时间", ""), "", "",
                row.get("备注", ""), "成功", "", captured_at,
            ])
        for index, row in enumerate(result.training_record_rows, start=1):
            rows.append([record.employee_id, name, "培训记录", row.get("来源页码", str(index)), "", "", row.get("培训科目", ""), "", row.get("培训机型", ""), row.get("培训时间", ""), row.get("培训結束时间", ""), "", "", "", "成功", "", captured_at])
        for index, row in enumerate(result.training_experience_rows, start=1):
            rows.append([record.employee_id, name, "训练经历", row.get("来源页码", str(index)), row.get("类型", ""), "", row.get("训练科目", ""), "", row.get("训练机型", ""), row.get("训练日期", ""), "", row.get("检查单", ""), "", row.get("结论", ""), "成功", "", captured_at])
        return rows

    def write_success(self, index: int, record: QueryRecord, result: QueryResult) -> None:
        del index
        workbook = load_workbook(self.paths.excel)
        try:
            for row in self._detail_rows(record, result):
                workbook["技术资料明细"].append(row)
            for key, value in result.basic_info.items():
                workbook["基础信息"].append([record.employee_id, result.page_name, "基本信息", 1, key, value])
            for section, section_rows in (("教育经历", result.education_rows), ("工作经历", result.work_rows), ("职称信息", result.title_rows), ("家庭信息", result.family_rows)):
                for row_index, row in enumerate(section_rows, start=1):
                    for key, value in row.items():
                        workbook["基础信息"].append([record.employee_id, result.page_name, section, row_index, key, value])
            for row in result.technical_rows:
                workbook["技术等级"].append([record.employee_id, result.page_name, *[row.get(header, "") for header in TECHNICAL_HEADERS]])
            for row in result.operation_rows:
                workbook["运行资格"].append([record.employee_id, result.page_name, *[row.get(header, "") for header in OPERATION_HEADERS]])
            for row in result.training_record_rows:
                workbook["培训记录"].append([record.employee_id, result.page_name, *[row.get(header, "") for header in TRAINING_RECORD_HEADERS], row.get("来源页码", "")])
            for row in result.training_experience_rows:
                workbook["训练经历"].append([record.employee_id, result.page_name, *[row.get(header, "") for header in TRAINING_EXPERIENCE_HEADERS], row.get("来源页码", "")])
            name_match = "未提供" if not record.name else (
                "是" if normalize_name(record.name) == normalize_name(result.page_name) else "否"
            )
            workbook["处理报告"].append([
                record.input_row,
                record.employee_id,
                record.name,
                result.page_name,
                "是",
                name_match,
                result.basic_count,
                len(result.technical_rows),
                len(result.operation_rows),
                len(result.training_record_rows),
                len(result.training_experience_rows),
                "成功",
                "",
                now_text(),
            ])
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()
        self._json_people.append({
            "employeeId": record.employee_id,
            "inputName": record.name,
            "status": "成功",
            "data": asdict(result),
        })
        self._write_json()

    def write_failure(self, index: int, record: QueryRecord, reason: str) -> None:
        del index
        workbook = load_workbook(self.paths.excel)
        try:
            workbook["处理报告"].append([
                record.input_row,
                record.employee_id,
                record.name,
                "",
                "否",
                "未查询",
                0,
                0,
                0,
                0,
                0,
                "失败",
                str(reason),
                now_text(),
            ])
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()
        self._json_people.append({"employeeId": record.employee_id, "inputName": record.name, "status": "失败", "error": str(reason)})
        self._write_json()

    def finalize(
        self,
        total: int,
        success: int,
        failed: int,
        input_errors: int,
        interrupted: bool,
        input_source: str = "",
    ) -> OutputPaths:
        summary_rows = [
            ("输入来源", input_source or "粘贴输入"),
            ("结果文件", str(self.paths.excel.resolve())),
            ("查询时间", now_text()),
            ("有效员工数", total),
            ("成功人数", success),
            ("失败人数", failed),
            ("输入错误数", input_errors),
            ("是否中断", "是" if interrupted else "否"),
        ]
        workbook = load_workbook(self.paths.excel)
        try:
            worksheet = workbook["汇总"]
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row - 1)
            for row in summary_rows:
                worksheet.append(row)
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()

        self._json_summary = {
            "inputSource": input_source or "粘贴输入",
            "queriedAt": now_text(),
            "total": total,
            "success": success,
            "failed": failed,
            "inputErrors": input_errors,
            "interrupted": interrupted,
        }
        lines = ["飞行人员信息查询报告", "=" * 48]
        lines.extend(f"{key}: {value}" for key, value in summary_rows)
        self.paths.report.write_text("\n".join(lines) + "\n", encoding="utf-8")
        self._write_json()
        return self.paths
