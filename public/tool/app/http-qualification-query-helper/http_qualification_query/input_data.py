from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from .models import InputIssue, InputPayload, QueryRecord


EMPLOYEE_HEADERS = {"员工号", "工号", "员工编号"}
NAME_HEADERS = {"姓名", "员工姓名"}
EMPLOYEE_PATTERN = re.compile(r"(?<!\d)(\d{6})(?!\d)")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_employee_id(value: Any) -> str:
    if value in (None, ""):
        return ""
    numeric_value = isinstance(value, (int, float)) and not isinstance(value, bool)
    if isinstance(value, float):
        if not value.is_integer():
            return ""
        text = str(int(value))
    elif isinstance(value, int):
        text = str(value)
    else:
        text = normalize_text(value)
        if re.fullmatch(r"\d+\.0+", text):
            text = text.split(".", 1)[0]
    if not text.isdigit() or len(text) > 6:
        return ""
    if numeric_value:
        return text.zfill(6)
    return text if len(text) == 6 else ""


def _find_header(headers: list[str], candidates: set[str]) -> int | None:
    return next((index for index, header in enumerate(headers) if header in candidates), None)


def read_excel_records(source: str | Path | BinaryIO) -> tuple[list[QueryRecord], list[InputIssue]]:
    workbook = load_workbook(source, data_only=True)
    records: list[QueryRecord] = []
    issues: list[InputIssue] = []
    seen: set[str] = set()
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            return [], [InputIssue(0, "", "", "Excel 为空", "Excel")]
        headers = [normalize_text(value) for value in rows[0]]
        employee_index = _find_header(headers, EMPLOYEE_HEADERS)
        name_index = _find_header(headers, NAME_HEADERS)
        if employee_index is None:
            return [], [InputIssue(1, "", "", "未找到员工号、工号或员工编号表头", "Excel 第1行")]

        for row_number, row in enumerate(rows[1:], start=2):
            if not row or not any(normalize_text(value) for value in row):
                continue
            raw_employee = row[employee_index] if employee_index < len(row) else ""
            employee_id = normalize_employee_id(raw_employee)
            name = normalize_text(row[name_index]) if name_index is not None and name_index < len(row) else ""
            row_source = f"Excel 第{row_number}行"
            if not employee_id:
                issues.append(InputIssue(row_number, normalize_text(raw_employee), name, "员工号不是六位数字", row_source))
                continue
            if employee_id in seen:
                issues.append(InputIssue(row_number, employee_id, name, "重复员工号，已跳过", row_source))
                continue
            seen.add(employee_id)
            records.append(QueryRecord(row_number, employee_id, name, row_source))
    finally:
        workbook.close()
    return records, issues


def parse_pasted_records(text: str) -> tuple[list[QueryRecord], list[InputIssue]]:
    records: list[QueryRecord] = []
    issues: list[InputIssue] = []
    seen: set[str] = set()
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    for line_number, line in enumerate(lines, start=1):
        source = f"粘贴第{line_number}条"
        employee_match = EMPLOYEE_PATTERN.search(line)
        if not employee_match:
            issues.append(InputIssue(line_number, "", "", "未识别六位员工号", source))
            continue
        employee_id = employee_match.group(1)
        remaining = f"{line[:employee_match.start()]} {line[employee_match.end():]}"
        name_match = re.search(r"[\u4e00-\u9fff·]{2,8}", remaining)
        name = name_match.group(0) if name_match else ""
        if employee_id in seen:
            issues.append(InputIssue(line_number, employee_id, name, "重复员工号，已跳过", source))
            continue
        seen.add(employee_id)
        records.append(QueryRecord(line_number, employee_id, name, source))
    return records, issues


def read_input(payload: InputPayload) -> tuple[list[QueryRecord], list[InputIssue]]:
    if payload.input_mode == "excel":
        if not payload.excel_bytes:
            raise ValueError("请选择 Excel 文件")
        suffix = Path(payload.excel_name).suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise ValueError("只支持 .xlsx 或 .xlsm 文件")
        return read_excel_records(BytesIO(payload.excel_bytes))
    if payload.input_mode == "paste":
        if not str(payload.pasted_text or "").strip():
            raise ValueError("请粘贴查询人员")
        return parse_pasted_records(payload.pasted_text)
    raise ValueError(f"不支持的输入方式：{payload.input_mode}")
