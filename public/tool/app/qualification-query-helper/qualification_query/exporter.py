from __future__ import annotations

import os
import re
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .models import InputIssue, OutputPaths, QueryRecord, QueryResult


DETAIL_HEADERS = (
    "员工号",
    "姓名",
    "资料类型",
    "页面序号",
    "类型",
    "代码",
    "名称",
    "水平等级",
    "机型",
    "生效时间",
    "失效时间",
    "对应检查记录",
    "数据来源",
    "备注",
    "抓取状态",
    "说明",
    "抓取时间",
)
REPORT_HEADERS = (
    "输入行号",
    "员工号",
    "输入姓名",
    "页面姓名",
    "员工号匹配",
    "姓名匹配",
    "技术等级条数",
    "运行资格条数",
    "抓取状态",
    "说明",
    "查询时间",
)
SUMMARY_HEADERS = ("项目", "值")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_name(value: str) -> str:
    text = re.sub(r"[（(][^）)]*[）)]", "", str(value or ""))
    return re.sub(r"\s+", "", text)


def _safe_run_name(run_id: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", run_id).strip("-") or "run"


def _save_atomic(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


class ResultExporter:
    def __init__(self, output_directory: Path, run_id: str):
        output_directory.mkdir(parents=True, exist_ok=True)
        name = _safe_run_name(run_id)
        self.paths = OutputPaths(
            excel=output_directory / f"技术等级运行资格查询_{name}.xlsx",
            report=output_directory / f"技术等级运行资格查询_{name}.txt",
        )

    def initialize(self, records: list[QueryRecord], input_issues: list[InputIssue]) -> None:
        workbook = Workbook()
        report_sheet = workbook.active
        report_sheet.title = "处理报告"
        detail_sheet = workbook.create_sheet("技术资料明细")
        summary_sheet = workbook.create_sheet("汇总")
        report_sheet.append(list(REPORT_HEADERS))
        detail_sheet.append(list(DETAIL_HEADERS))
        summary_sheet.append(list(SUMMARY_HEADERS))
        for worksheet in (report_sheet, detail_sheet, summary_sheet):
            _style_header(worksheet)

        for column, width in {
            "A": 11, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
            "G": 16, "H": 16, "I": 14, "J": 48, "K": 20,
        }.items():
            report_sheet.column_dimensions[column].width = width
        for column, width in {
            "A": 12, "B": 12, "C": 12, "D": 10, "E": 18, "F": 16, "G": 34,
            "H": 12, "I": 10, "J": 14, "K": 14, "L": 18, "M": 12, "N": 12,
            "O": 12, "P": 48, "Q": 20,
        }.items():
            detail_sheet.column_dimensions[column].width = width
        summary_sheet.column_dimensions["A"].width = 20
        summary_sheet.column_dimensions["B"].width = 90

        for issue in input_issues:
            report_sheet.append([
                issue.input_row,
                issue.employee_id,
                issue.name,
                "",
                "否",
                "未查询",
                0,
                0,
                "输入错误",
                issue.message,
                now_text(),
            ])
        _save_atomic(workbook, self.paths.excel)
        workbook.close()

    @staticmethod
    def _detail_rows(record: QueryRecord, result: QueryResult) -> list[list[str]]:
        name = result.page_name or record.name
        captured_at = now_text()
        rows: list[list[str]] = []
        for row in result.technical_rows:
            rows.append([
                record.employee_id, name, "技术等级", row.get("#", ""), "", row.get("技术等级代码", ""),
                row.get("技术等级", ""), row.get("水平等级", ""), row.get("机型", ""),
                row.get("生效时间", ""), row.get("失效时间", ""), row.get("对应检查记录", ""),
                row.get("数据来源", ""), "", "成功", "", captured_at,
            ])
        for index, row in enumerate(result.operation_rows, start=1):
            rows.append([
                record.employee_id, name, "运行资格", str(index), row.get("类型", ""), row.get("运行资格代码", ""),
                row.get("运行资格", ""), row.get("水平等级", ""), row.get("机型", ""),
                row.get("生效时间", ""), row.get("失效时间", ""), "", "",
                row.get("备注", ""), "成功", "", captured_at,
            ])
        return rows

    def write_success(self, index: int, record: QueryRecord, result: QueryResult) -> None:
        del index
        workbook = load_workbook(self.paths.excel)
        try:
            for row in self._detail_rows(record, result):
                workbook["技术资料明细"].append(row)
            name_match = "未提供" if not record.name else (
                "是" if normalize_name(record.name) == normalize_name(result.page_name) else "否"
            )
            workbook["处理报告"].append([
                record.input_row,
                record.employee_id,
                record.name,
                result.page_name,
                "是",
                name_match,
                len(result.technical_rows),
                len(result.operation_rows),
                "成功",
                "",
                now_text(),
            ])
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()

    def write_failure(self, index: int, record: QueryRecord, reason: str) -> None:
        del index
        workbook = load_workbook(self.paths.excel)
        try:
            workbook["处理报告"].append([
                record.input_row,
                record.employee_id,
                record.name,
                "",
                "否",
                "未查询",
                0,
                0,
                "失败",
                str(reason),
                now_text(),
            ])
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()

    def finalize(
        self,
        total: int,
        success: int,
        failed: int,
        input_errors: int,
        interrupted: bool,
        input_source: str = "",
    ) -> OutputPaths:
        summary_rows = [
            ("输入来源", input_source or "粘贴输入"),
            ("结果文件", str(self.paths.excel.resolve())),
            ("查询时间", now_text()),
            ("有效员工数", total),
            ("成功人数", success),
            ("失败人数", failed),
            ("输入错误数", input_errors),
            ("是否中断", "是" if interrupted else "否"),
        ]
        workbook = load_workbook(self.paths.excel)
        try:
            worksheet = workbook["汇总"]
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row - 1)
            for row in summary_rows:
                worksheet.append(row)
            _save_atomic(workbook, self.paths.excel)
        finally:
            workbook.close()

        lines = ["技术等级运行资格查询报告", "=" * 48]
        lines.extend(f"{key}: {value}" for key, value in summary_rows)
        self.paths.report.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return self.paths
