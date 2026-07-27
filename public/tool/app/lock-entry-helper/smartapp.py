# 锁班皇帝 - 智能路由助手

import re
import platform
import os
from datetime import date, datetime, timedelta
from colorama import init, Fore, Style
from playwright.sync_api import sync_playwright

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.datetime import from_excel
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

init()  # 初始化colorama


def beep_error():
    """错误提示音"""
    try:
        if platform.system() == 'Windows':
            import winsound
            winsound.Beep(800, 300)
        else:
            os.system('paplay /usr/share/sounds/freedesktop/stereo/dialog-error.oga 2>/dev/null || true')
    except:
        pass

# 所有原始锁班类型均可录入，仅两类休假参与智能路由。
LEAVE_TYPE_MAP = {
    "ALV-年假（公休假）": "ALV",
    "ALV_FD-飞行员公休（订座）": "ALV_FD",
    "RECU_LVE-健康疗养": "RECU_LVE",
    "RECU_LVE_R-康复疗养": "RECU_LVE_R",
    "MAT_FA_LVE-陪产假": "MAT_FA_LVE",
    "MAT_MO_LVE-产假": "MAT_MO_LVE",
    "PREGNANT-孕假": "PREGNANT",
    "PARENT_LVE-探亲假-探父母": "PARENT_LVE",
    "SPOUSE_LVE-探亲假-探配偶": "SPOUSE_LVE",
    "MARR_LVE-婚假": "MARR_LVE",
    "COMP_LVE-丧假": "COMP_LVE",
    "CHILD_LVE-育儿假": "CHILD_LVE",
    "INJURY_LVE-工伤假": "INJURY_LVE",
    "LWOP_LVE-其他（事假）": "LWOP_LVE",
    "UNPAID_LVE-无薪": "UNPAID_LVE",
    "HOUSE_LVE-搬家": "HOUSE_LVE",
    "BREED_LVE-哺乳假": "BREED_LVE",
    "PATERNITY-独生子女护理假": "PATERNITY",
    "BIRC_LVE-计划生育假": "BIRC_LVE",
    "REWARD_LVE-奖励": "REWARD_LVE",
    "PENALTY-停飞": "PENALTY",
    "PRD_LVE-经期假": "PRD_LVE",
    "GRD-地面班": "GRD",
    "GDO-地面休息": "GDO",
    "TRNG1-训练": "TRNG1",
    "BS_STUDY-业务学习": "BS_STUDY",
    "BUSINESS-公务": "BUSINESS",
    "GRD_ONDUTY-地面值班": "GRD_ONDUTY",
    "LG_STUDY-语言学习/考试": "LG_STUDY",
    "MEDL_CHK-体检_临床": "MEDL_CHK",
    "MEDL_PHLE-体检_抽血": "MEDL_PHLE",
    "MEDL_EET-体检_平板": "MEDL_EET",
    "MEDL_PSYC-体检_心理测试": "MEDL_PSYC",
    "MTG-会议": "MTG",
    "MTG_SF-安全讲评会": "MTG_SF",
    "DGET-危险品培训": "DGET",
    "EP-飞行人员应急复训": "EP",
    "CRM-CRM培训": "CRM",
    "T_SIM_INS-模拟机检查": "T_SIM_INS",
    "T_SIM_REC-模拟机复训": "T_SIM_REC",
    "T_SIM_INT-模拟机初始": "T_SIM_INT",
    "T_SIM_UPG-模拟机升级": "T_SIM_UPG",
    "T_SIM_CON-模拟机_转机型": "T_SIM_CON",
    "MAKEUP-补考": "MAKEUP",
    "BS_CONCL-飞行后讲评": "BS_CONCL",
    "BS_CHK-业务检查": "BS_CHK",
    "ADMN-管理任务": "ADMN",
    "SOCIAL-社会活动": "SOCIAL",
    "HANDBOOK-手册": "HANDBOOK",
    "POL_STUDY-政治学习": "POL_STUDY",
    "T/A-部门活动": "T/A",
}
SMART_LEAVE_TYPES = ("RECU_LVE", "ALV_FD")
ALTERNATE_LEAVE_TYPE = {
    "RECU_LVE": "ALV_FD",
    "ALV_FD": "RECU_LVE",
}

# 代码到中文名的反向映射
LEAVE_CODE_TO_NAME = {v: k for k, v in LEAVE_TYPE_MAP.items()}
LEAVE_CODE_PATTERN = re.compile(
    r'(?<![A-Z0-9_/])('
    + '|'.join(re.escape(code) for code in sorted(LEAVE_CODE_TO_NAME, key=len, reverse=True))
    + r')(?![A-Z0-9_/])'
)

RESULT_HEADERS = [
    "原始序号",
    "片段序号",
    "员工号",
    "姓名",
    "输入锁班类型",
    "输入开始日期",
    "输入结束日期",
    "实际锁班类型",
    "实际开始日期",
    "实际结束日期",
    "计划天数",
    "输入类型可休天数",
    "替代类型可休天数",
    "处理状态",
    "锁班结果",
    "结果姓名",
    "结果锁班类型",
    "结果开始日期",
    "结果结束日期",
    "冲突",
    "备注",
    "员工号匹配",
    "姓名匹配",
    "日期匹配",
    "类型匹配",
    "处理时间",
    "尝试次数",
    "冲突回退",
    "解锁序号",
    "解锁状态",
    "解锁员工号",
    "解锁姓名",
    "解锁开始日期",
    "解锁结束日期",
    "解锁天数",
    "解锁类型",
    "解锁名称",
    "解锁原因",
    "解锁录入人",
    "解锁录入时间",
]

QUOTA_REQUIRED_HEADERS = [
    "休假类型",
    "年份",
    "休假天数",
    "锁班天数",
    "解锁天数",
    "可休天数",
]

EXCEL_HEADER_ALIASES = {
    "员工号": ("员工号", "工号"),
    "姓名": ("姓名",),
    "锁班类型": ("锁班类型", "请假类型"),
    "开始日期": ("开始日期",),
    "结束日期": ("结束日期",),
}

LOCK_QUERY_REQUIRED_HEADERS = [
    "序号",
    "状态",
    "员工号",
    "姓名",
    "开始日期",
    "结束日期",
    "锁班天数",
    "锁班类型",
    "锁班名称",
    "锁班原因",
    "录入人",
    "录入时间",
]

def parse_leave_type(text: str) -> str:
    """解析锁班类型，支持代码或中文名"""
    if not text:
        return None
    text = str(text).strip()
    # 直接是代码
    if text in LEAVE_CODE_TO_NAME:
        return text
    # 完整格式 "CODE-中文名"
    if text in LEAVE_TYPE_MAP:
        return LEAVE_TYPE_MAP[text]
    # 从整行文本里提取代码。
    code_match = LEAVE_CODE_PATTERN.search(text)
    if code_match:
        return code_match.group(1)
    # 从整行文本里提取完整格式，优先匹配更长的名称
    for key, code in sorted(LEAVE_TYPE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        if key in text:
            return code
    # 只有中文名，模糊匹配
    for key, code in sorted(LEAVE_TYPE_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        name_part = key.split('-', 1)[1] if '-' in key else key
        if name_part in text or text in name_part:
            return code
    return None


def normalize_employee_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_portal_business_date(value: str) -> date:
    """从门户日期或日期时间文本读取业务日期。"""
    text = normalize_text(value)
    match = re.match(r"^(\d{4}-\d{2}-\d{2})(?:\s|$)", text)
    if not match:
        raise ValueError(f"门户日期格式异常 [{text}]")
    return datetime.strptime(match.group(1), "%Y-%m-%d").date()


def date_ranges_overlap(
    first_start: str,
    first_end: str,
    second_start: str,
    second_end: str,
) -> bool:
    """按包含首尾的自然日区间判断两条锁班是否重叠。"""
    first_start_date = parse_portal_business_date(first_start)
    first_end_date = parse_portal_business_date(first_end)
    second_start_date = parse_portal_business_date(second_start)
    second_end_date = parse_portal_business_date(second_end)
    if first_end_date < first_start_date or second_end_date < second_start_date:
        raise ValueError("锁班结束日期早于开始日期")
    return first_start_date <= second_end_date and second_start_date <= first_end_date


def choose_unlock_candidate(rows: list, record: dict) -> tuple:
    """只接受员工号、已锁状态和日期区间共同唯一命中的旧记录。"""
    employee_id = normalize_employee_id(record.get("员工号"))
    try:
        parse_portal_business_date(record.get("开始日期", ""))
        parse_portal_business_date(record.get("结束日期", ""))
    except ValueError as error:
        return None, f"冲突记录日期异常: {error}"

    candidates = []
    for row in rows:
        if normalize_employee_id(row.get("员工号")) != employee_id:
            continue
        if normalize_text(row.get("状态")) != "已锁":
            continue
        try:
            overlaps = date_ranges_overlap(
                row.get("开始日期", ""),
                row.get("结束日期", ""),
                record.get("开始日期", ""),
                record.get("结束日期", ""),
            )
        except ValueError as error:
            return None, f"已锁候选日期异常: {error}"
        if overlaps:
            candidates.append(row)

    if not candidates:
        return None, "未找到员工号一致、状态已锁且日期重叠的旧锁班"
    if len(candidates) != 1:
        return None, f"找到{len(candidates)}条日期重叠的已锁记录，无法唯一定位"
    return candidates[0], ""


def format_unlocked_record_excel_note(row: dict) -> str:
    """生成结果 Excel 使用的旧锁班摘要。"""
    return (
        f"已解锁：锁班名称{normalize_text(row.get('锁班名称'))}；"
        f"锁班原因{normalize_text(row.get('锁班原因'))}；"
        f"开始日期{normalize_text(row.get('开始日期'))}；"
        f"结束日期{normalize_text(row.get('结束日期'))}"
    )


def format_business_date(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, int) and re.fullmatch(r"(?:19|20)\d{6}", str(value)):
            return normalize_date(str(value))
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except (TypeError, ValueError, OverflowError):
            return ""
    if isinstance(value, str):
        return normalize_date(value)
    return ""


def excel_header_indexes(header_row: tuple) -> tuple[dict, list]:
    normalized = [normalize_text(value) for value in header_row]
    indexes = {}
    missing = []
    for field, aliases in EXCEL_HEADER_ALIASES.items():
        index = next((normalized.index(alias) for alias in aliases if alias in normalized), None)
        if index is None:
            missing.append(field)
        else:
            indexes[field] = index
    return indexes, missing


def parse_excel_file(filepath: str, whitelist: set = None) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    if not HAS_OPENPYXL:
        return [], ["未安装openpyxl库，请运行: pip install openpyxl"]

    records = []
    errors = []

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        indexes, missing_headers = excel_header_indexes(header_row)
        if missing_headers:
            wb.close()
            return [], [f"Excel缺少表头: {', '.join(missing_headers)}"]

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            emp_id = normalize_employee_id(row[indexes["员工号"]])
            name = normalize_text(row[indexes["姓名"]]) or None
            leave_type_raw = normalize_text(row[indexes["锁班类型"]])
            start = format_business_date(row[indexes["开始日期"]])
            end = format_business_date(row[indexes["结束日期"]]) or start

            leave_type = parse_leave_type(leave_type_raw)
            record = {
                "员工号": emp_id,
                "姓名": name,
                "请假类型": leave_type,
                "开始日期": start,
                "结束日期": end,
            }
            problem = validate_record(record)
            if problem:
                errors.append(f"第{row_num}行: {problem}")
                continue
            if whitelist and emp_id not in whitelist:
                continue
            records.append(record)

        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")

    return records, errors


def c_info(text):
    return f"{Fore.CYAN}{text}{Style.RESET_ALL}"

def c_ok(text):
    return f"{Fore.GREEN}{text}{Style.RESET_ALL}"

def c_err(text):
    return f"{Fore.RED}{text}{Style.RESET_ALL}"

def c_warn(text):
    return f"{Fore.YELLOW}{text}{Style.RESET_ALL}"

def c_hint(text):
    return f"{Fore.MAGENTA}{text}{Style.RESET_ALL}"


def parse_whitelist(text: str) -> set:
    """解析员工号白名单"""
    all_nums = re.findall(r'\d{6}', re.sub(r'\D', ' ', text))
    if all_nums:
        return set(all_nums)
    text = re.sub(r'\D', '', text)
    return set(text[i:i+6] for i in range(0, len(text), 6) if len(text[i:i+6]) == 6)


def normalize_text(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_date(date_str) -> str:
    """把各种日期格式统一转成YYYY-MM-DD"""
    if date_str is None:
        return ""
    if isinstance(date_str, (datetime, date)):
        return date_str.strftime('%Y-%m-%d')
    date_str = str(date_str).strip()
    compact_date = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", date_str)
    if compact_date:
        return "-".join(compact_date.groups())
    separated_date = re.fullmatch(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", date_str)
    if separated_date:
        year, month, day = separated_date.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return date_str


def parse_iso_date(value, label: str = "日期") -> date:
    normalized = normalize_date(value)
    try:
        return datetime.strptime(normalized, "%Y-%m-%d").date()
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label}格式错误 [{value or ''}]") from error


def validate_record(record: dict) -> str:
    employee_id = normalize_employee_id(record.get("员工号"))
    if not re.fullmatch(r"\d{6}", employee_id):
        return f"员工号格式错误 [{employee_id}]"
    leave_type = record.get("请假类型")
    if leave_type not in LEAVE_CODE_TO_NAME:
        return "未识别锁班类型"
    try:
        start = parse_iso_date(record.get("开始日期"), "开始日期")
        end = parse_iso_date(record.get("结束日期"), "结束日期")
    except ValueError as error:
        return str(error)
    if end < start:
        return "结束日期不得早于开始日期"
    if leave_type in SMART_LEAVE_TYPES and start.year != end.year:
        return "智能路由暂不处理跨自然年记录"
    return ""


def leave_type_name(leave_type: str) -> str:
    display = LEAVE_CODE_TO_NAME.get(leave_type, leave_type or "")
    return display.split('-', 1)[1] if '-' in display else display


def normalize_person_name(name: str) -> str:
    text = normalize_text(name).upper()
    text = re.sub(r'[（(][^）)]*[）)]', '', text)
    return re.sub(r'\s+', ' ', text).strip()


def names_match(expected: str, actual: str) -> bool:
    if not expected or not actual:
        return True
    return normalize_person_name(expected) == normalize_person_name(actual)


def leave_types_match(expected_code: str, actual_text: str) -> bool:
    if not actual_text:
        return True
    actual_code = parse_leave_type(actual_text)
    if actual_code:
        return actual_code == expected_code
    expected_name = leave_type_name(expected_code)
    return bool(expected_name and expected_name in normalize_text(actual_text))


def same_day(left, right) -> bool:
    return normalize_date(str(left or '')[:10]) == normalize_date(str(right or '')[:10])


def parse_quota_rows(headers: list, rows: list) -> list:
    normalized_headers = [normalize_text(value) for value in headers]
    missing = [header for header in QUOTA_REQUIRED_HEADERS if header not in normalized_headers]
    if missing:
        raise ValueError(f"休假限制表缺少表头: {', '.join(missing)}")

    parsed = []
    for row_index, values in enumerate(rows, start=1):
        normalized_values = [normalize_text(value) for value in values]
        if not any(normalized_values):
            continue
        if len(normalized_values) != len(normalized_headers):
            raise ValueError(
                f"休假限制表第{row_index}行列数异常: 表头{len(normalized_headers)}列，数据{len(normalized_values)}列"
            )
        parsed.append(dict(zip(normalized_headers, normalized_values)))
    return parsed


def parse_available_days(value) -> int:
    if isinstance(value, bool):
        raise ValueError(f"可休天数不是非负整数 [{value}]")
    if isinstance(value, int):
        if value < 0:
            raise ValueError(f"可休天数不是非负整数 [{value}]")
        return value
    text = normalize_text(value)
    if not re.fullmatch(r"\d+", text):
        raise ValueError(f"可休天数不是非负整数 [{text}]")
    return int(text)


def available_days_for_year(rows: list, year: int) -> int:
    matches = [row for row in rows if normalize_text(row.get("年份")) == str(year)]
    if not matches:
        raise ValueError(f"休假限制表没有{year}年数据")
    if len(matches) > 1:
        raise ValueError(f"休假限制表存在多条{year}年数据")
    return parse_available_days(matches[0].get("可休天数"))


def route_record(record: dict, available_by_type: dict) -> tuple:
    problem = validate_record(record)
    if problem:
        return [], problem

    primary_type = record["请假类型"]
    start = parse_iso_date(record["开始日期"], "开始日期")
    end = parse_iso_date(record["结束日期"], "结束日期")
    requested_days = (end - start).days + 1
    if primary_type not in SMART_LEAVE_TYPES:
        return [{
            "员工号": record.get("员工号"),
            "姓名": record.get("姓名"),
            "请假类型": primary_type,
            "开始日期": start.strftime("%Y-%m-%d"),
            "结束日期": end.strftime("%Y-%m-%d"),
            "计划天数": requested_days,
        }], ""

    alternate_type = ALTERNATE_LEAVE_TYPE[primary_type]
    try:
        primary_available = parse_available_days(available_by_type.get(primary_type, 0))
        alternate_available = parse_available_days(available_by_type.get(alternate_type, 0))
    except ValueError as error:
        return [], str(error)

    combined_available = primary_available + alternate_available
    if combined_available < requested_days:
        return [], (
            f"两类可休天数合计{combined_available}天，少于需要{requested_days}天；"
            f"{leave_type_name(primary_type)}{primary_available}天，"
            f"{leave_type_name(alternate_type)}{alternate_available}天"
        )

    primary_days = min(primary_available, requested_days)
    alternate_days = requested_days - primary_days
    segments = []
    cursor = start
    for leave_type, days in (
        (primary_type, primary_days),
        (alternate_type, alternate_days),
    ):
        if days <= 0:
            continue
        segment_end = cursor + timedelta(days=days - 1)
        segments.append({
            "员工号": record.get("员工号"),
            "姓名": record.get("姓名"),
            "请假类型": leave_type,
            "开始日期": cursor.strftime("%Y-%m-%d"),
            "结束日期": segment_end.strftime("%Y-%m-%d"),
            "计划天数": days,
        })
        cursor = segment_end + timedelta(days=1)
    return segments, ""


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {"员工号": None, "姓名": None, "请假类型": None, "开始日期": None, "结束日期": None}
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    result["请假类型"] = parse_leave_type(text)
    dates = re.findall(r'\d{8}|\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', text)
    if dates:
        result["开始日期"] = normalize_date(dates[0])
        result["结束日期"] = normalize_date(dates[1]) if len(dates) > 1 else normalize_date(dates[0])
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    # 按6位员工号切分
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str, whitelist: set = None) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    # 先按换行分，如果只有一行且很长，尝试按员工号切分
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        problem = validate_record(record)
        if problem:
            errors.append(f"第{i}条: {problem} [{line[:50]}]")
            continue
        if whitelist and record["员工号"] not in whitelist:
            continue
        records.append(record)
    return records, errors


def clear_form(page):
    """清空表单"""
    page.locator("#showIdshowNonproductionTaskImportPage").fill("")
    page.locator("#lockStartTime").fill("")
    page.locator("#lockEndTime").fill("")
    clear_reason_field(page)


def clear_reason_field(page):
    """尽量清空备注框，避免沿用上一条记录的备注"""
    reason_input = page.locator("#lockReasonTxt")
    if reason_input.count() == 0:
        return
    try:
        reason_input.scroll_into_view_if_needed()
        reason_input.click()
        page.wait_for_timeout(100)
        reason_input.fill("")
        page.wait_for_timeout(100)
    except Exception:
        pass


def fill_reason_field(page, reason_text):
    """按门户校验要求填写备注"""
    if not reason_text:
        return
    try:
        print(c_info("正在填写备注..."))
        reason_input = page.locator("#lockReasonTxt")
        reason_input.wait_for(timeout=5000)
        reason_input.scroll_into_view_if_needed()
        reason_input.click()
        page.wait_for_timeout(100)
        reason_input.fill("")
        page.wait_for_timeout(100)
        reason_input.click()
        page.wait_for_timeout(100)
        reason_input.type(reason_text, delay=120)
    except Exception as error:
        raise RuntimeError(f"填写备注失败: {error}") from error


def fill_employee(page, emp_id, expected_name=None):
    emp_input = page.locator("#showIdshowNonproductionTaskImportPage")
    emp_input.click()
    emp_input.fill("")
    emp_input.type(str(emp_id), delay=10)
    page.evaluate("""
        const input = document.querySelector('#showIdshowNonproductionTaskImportPage');
        if (input) {
            input.dispatchEvent(new Event('blur', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
        }
    """)
    expected_normalized_name = normalize_person_name(expected_name) if expected_name else ""
    page.wait_for_function(
        r"""expected => {
            const hidden = document.querySelector('#nonproductionTaskImportStaffNumId');
            const name = document.querySelector('#nameInfo');
            const normalizeName = value => String(value || '')
                .toUpperCase()
                .replace(/[（(][^）)]*[）)]/g, '')
                .replace(/\s+/g, ' ')
                .trim();
            return hidden
                && hidden.value === expected.employeeId
                && name
                && name.value.trim()
                && (!expected.name || normalizeName(name.value) === expected.name);
        }""",
        arg={"employeeId": str(emp_id), "name": expected_normalized_name},
        timeout=15000,
    )
    result_employee_id = page.locator("#nonproductionTaskImportStaffNumId").input_value()
    result_name = page.locator("#nameInfo").input_value()
    if result_employee_id != str(emp_id):
        raise RuntimeError(f"页面员工号识别不一致: 输入{emp_id}，页面{result_employee_id}")
    if expected_name and not names_match(expected_name, result_name):
        raise RuntimeError(f"页面姓名识别不一致: 输入{expected_name}，页面{result_name}")
    return result_name


def select_leave_type(page, leave_type):
    if leave_type not in LEAVE_CODE_TO_NAME:
        raise ValueError(f"不支持锁班类型 [{leave_type}]")
    page.evaluate("""(leaveType) => {
        const select = document.querySelector('#showNonproductionTaskImportPage #lockType');
        if (select) {
            select.focus();
            select.value = leaveType;
            select.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
            select.dispatchEvent(new MouseEvent('click', { bubbles: true }));
            select.dispatchEvent(new Event('input', { bubbles: true }));
            select.dispatchEvent(new Event('change', { bubbles: true }));
            select.dispatchEvent(new Event('blur', { bubbles: true }));
        }
    }""", leave_type)
    page.wait_for_function(
        "leaveType => document.querySelector('#showNonproductionTaskImportPage #lockType')?.value === leaveType",
        arg=leave_type,
        timeout=5000,
    )


def fill_dates(page, start_date, end_date):
    start_input = page.locator("#lockStartTime")
    end_input = page.locator("#lockEndTime")
    start_input.fill(start_date)
    start_input.dispatch_event("change")
    end_input.fill(end_date)
    end_input.dispatch_event("change")
    end_input.dispatch_event("blur")


def fill_form(page, emp_id, leave_type, start_date, end_date, reason_text=None, expected_name=None):
    """填写一个实际路由片段。"""
    clear_form(page)
    fill_employee(page, emp_id, expected_name)
    select_leave_type(page, leave_type)
    fill_dates(page, start_date, end_date)
    fill_reason_field(page, reason_text)


def is_row_number(value: str) -> bool:
    return bool(re.fullmatch(r'\d{1,4}', normalize_text(value)))


def align_table_values(headers: list, values: list) -> list:
    if headers and values and len(values) == len(headers) + 1 and is_row_number(values[0]) and headers[0] != "序号":
        return values[1:]
    return values


def table_headers(page, selector: str) -> list:
    cells = page.locator(selector)
    headers = []
    for index in range(cells.count()):
        try:
            text = normalize_text(cells.nth(index).inner_text(timeout=1000))
        except Exception:
            text = ""
        if text:
            headers.append(text)
    return headers


def table_rows(page, selector: str, headers: list) -> list:
    row_locs = page.locator(selector)
    rows = []
    for row_index in range(row_locs.count()):
        row = row_locs.nth(row_index)
        cell_locs = row.locator("td")
        values = []
        for cell_index in range(cell_locs.count()):
            try:
                values.append(normalize_text(cell_locs.nth(cell_index).inner_text(timeout=1000)))
            except Exception:
                values.append("")
        if not values:
            try:
                values = [normalize_text(row.inner_text(timeout=1000))]
            except Exception:
                values = []
        else:
            values = align_table_values(headers, values)
        row_map = {}
        for index, value in enumerate(values):
            key = headers[index] if index < len(headers) else f"列{index + 1}"
            row_map[key] = value
        row_map["_text"] = " | ".join(value for value in values if value)
        rows.append(row_map)
    return rows


def lock_query_row_from_locator(row, headers: list) -> dict:
    cells = row.locator("td")
    if cells.count() != len(headers):
        text = normalize_text(row.inner_text(timeout=1000))
        if "没有相关信息" in text:
            return {"_text": text}
        raise RuntimeError(
            f"锁班查询表数据列数异常: 表头{len(headers)}列，数据{cells.count()}列"
        )
    values = []
    for cell_index in range(cells.count()):
        cell = cells.nth(cell_index)
        title = normalize_text(cell.get_attribute("title") or "")
        text = normalize_text(cell.inner_text(timeout=1000))
        values.append(title or text)
    result = {headers[index]: value for index, value in enumerate(values)}
    result["_text"] = " | ".join(value for value in values if value)
    return result


def read_lock_query_page_rows(page) -> list:
    table = page.locator("#showNonproductionTaskResultPage table")
    table.wait_for(state="visible", timeout=15000)
    header_cells = table.locator("thead th")
    headers = []
    for index in range(header_cells.count()):
        text = normalize_text(header_cells.nth(index).inner_text(timeout=1000))
        headers.append(text or ("选择" if index == 0 else f"列{index + 1}"))
    missing = [header for header in LOCK_QUERY_REQUIRED_HEADERS if header not in headers]
    if missing:
        raise RuntimeError(f"锁班查询表缺少表头: {', '.join(missing)}")

    page_input = page.locator("#showLockListResultPageDiv #userPage")
    page_number = int(page_input.input_value()) if page_input.count() else 1
    rows = []
    row_locators = table.locator("tbody.list tr")
    for row_index in range(row_locators.count()):
        row = lock_query_row_from_locator(row_locators.nth(row_index), headers)
        if "没有相关信息" in row.get("_text", ""):
            continue
        row["_page_number"] = page_number
        rows.append(row)
    return rows


def lock_query_total_pages(page) -> int:
    links = page.locator("#showLockListResultPageDiv .footer a")
    for index in range(links.count()):
        link = links.nth(index)
        if normalize_text(link.inner_text(timeout=1000)) != "最后一页":
            continue
        href = link.get_attribute("href") or ""
        match = re.search(r",\s*['\"]?(\d+)['\"]?\);?$", href)
        if not match:
            raise RuntimeError(f"无法识别锁班查询总页数 [{href}]")
        return int(match.group(1))
    return 1


def go_to_lock_query_page(page, page_number: int) -> None:
    current = page.locator("#showLockListResultPageDiv #userPage")
    if current.count() and current.input_value() == str(page_number):
        return
    with page.expect_response(
        lambda response: "/newieb/nonproductionTask/showLockListPage" in response.url,
        timeout=15000,
    ):
        page.evaluate(
            """pageNumber => window.goPageTwo(
                'showNonproductionTaskLockPage1',
                'queryFormId',
                'showLockListResultPageDiv',
                String(pageNumber)
            )""",
            page_number,
        )
    page.wait_for_function(
        "pageNumber => document.querySelector('#showLockListResultPageDiv #userPage')?.value === String(pageNumber)",
        arg=page_number,
        timeout=15000,
    )


def read_all_lock_query_rows(page) -> list:
    go_to_lock_query_page(page, 1)
    total_pages = lock_query_total_pages(page)
    rows = []
    for page_number in range(1, total_pages + 1):
        go_to_lock_query_page(page, page_number)
        rows.extend(read_lock_query_page_rows(page))
    return rows


def run_locked_record_query(page, employee_id: str) -> list:
    query_page = page.locator("#showNonproductionTaskLockPage1")
    query_page.wait_for(state="visible", timeout=15000)
    employee = query_page.locator("#showIdNonproductionTaskLock")
    employee.click()
    employee.fill("")
    employee.type(employee_id, delay=20)
    page.wait_for_function(
        "employeeId => document.querySelector('#staffnumNonproductionTaskLock')?.value === employeeId",
        arg=employee_id,
        timeout=15000,
    )
    query_page.locator("#lockStatus").select_option("1")
    query_page.locator("#lockStartTimeQuery").fill("")
    query_page.locator("#lockEndTimeQuery").fill("")
    with page.expect_response(
        lambda response: "/newieb/nonproductionTask/showLockListPage" in response.url,
        timeout=15000,
    ):
        query_page.get_by_role("button", name="查询", exact=True).click()
    page.locator("#showNonproductionTaskResultPage table").wait_for(state="visible", timeout=15000)
    return read_all_lock_query_rows(page)


def open_lock_query_from_conflict(page, record: dict) -> list:
    page.get_by_role("button", name="锁班查询", exact=True).click()
    page.locator("#showNonproductionTaskLockPage1").wait_for(state="visible", timeout=15000)
    return run_locked_record_query(page, record["员工号"])


def same_query_record(left: dict, right: dict) -> bool:
    identity_fields = (
        "员工号",
        "状态",
        "开始日期",
        "结束日期",
        "锁班类型",
        "锁班名称",
        "锁班原因",
        "录入人",
        "录入时间",
    )
    return all(normalize_text(left.get(field)) == normalize_text(right.get(field)) for field in identity_fields)


def find_unlock_row_locator(page, candidate: dict):
    go_to_lock_query_page(page, int(candidate.get("_page_number", 1)))
    table = page.locator("#showNonproductionTaskResultPage table")
    header_cells = table.locator("thead th")
    headers = []
    for index in range(header_cells.count()):
        text = normalize_text(header_cells.nth(index).inner_text(timeout=1000))
        headers.append(text or ("选择" if index == 0 else f"列{index + 1}"))

    matches = []
    row_locators = table.locator("tbody.list tr")
    for row_index in range(row_locators.count()):
        row_locator = row_locators.nth(row_index)
        row = lock_query_row_from_locator(row_locator, headers)
        if same_query_record(row, candidate):
            matches.append(row_locator)
    if len(matches) != 1:
        raise RuntimeError(f"解锁前同行复选框定位到{len(matches)}行，已停止")
    return matches[0]


def format_unlock_action_reason(candidate: dict, record: dict) -> str:
    old_start = parse_portal_business_date(candidate.get("开始日期", "")).strftime("%Y-%m-%d")
    old_end = parse_portal_business_date(candidate.get("结束日期", "")).strftime("%Y-%m-%d")
    return (
        f"自动冲突回退：解锁{candidate.get('序号', '')} "
        f"{candidate.get('锁班类型', '')} {old_start}~{old_end}，"
        f"重提{record.get('请假类型', '')} {record.get('开始日期', '')}~{record.get('结束日期', '')}"
    )[:200]


def unlock_query_candidate(page, candidate: dict, record: dict) -> str:
    row_locator = find_unlock_row_locator(page, candidate)
    checkbox = row_locator.locator("input.lockTaskListIds")
    if checkbox.count() != 1 or checkbox.is_disabled():
        raise RuntimeError("唯一候选行的复选框不可用，已停止")
    checkbox.check()
    page.locator("#showNonproductionTaskResultPage #importBtn").get_by_role(
        "button", name="解锁", exact=True
    ).click()

    remark = page.locator("#remark")
    remark.wait_for(state="visible", timeout=10000)
    remark_dialog = remark.locator(
        "xpath=ancestor::div[contains(concat(' ', normalize-space(@class), ' '), ' ui-dialog ')][1]"
    )
    operator_prefix = normalize_text(remark.input_value())
    action_reason = format_unlock_action_reason(candidate, record)
    remark.fill((operator_prefix + action_reason)[:200])
    print(c_info("已填写解锁原因，确认解锁"))
    remark_dialog.get_by_role("button", name="确定", exact=True).click()

    remark.wait_for(state="hidden", timeout=10000)
    page.locator(".ui-dialog:visible").last.wait_for(state="visible", timeout=15000)
    visible_dialogs = page.locator(".ui-dialog:visible")
    result_dialog = visible_dialogs.nth(visible_dialogs.count() - 1)
    result_message = normalize_text(result_dialog.inner_text())
    print(c_info(f"解锁结果: {result_message}"))
    result_dialog.get_by_role("button", name="确定", exact=True).click()
    if "成功" not in result_message:
        raise RuntimeError(f"门户未确认解锁成功: {result_message}")
    return result_message


def return_to_import_from_query(page) -> None:
    if page.locator("#showIdshowNonproductionTaskImportPage").is_visible():
        return
    page.locator("#showNonproductionTaskResultPage #importBtn").get_by_role(
        "button", name="录入", exact=True
    ).click()
    page.locator("#showIdshowNonproductionTaskImportPage").wait_for(state="visible", timeout=15000)


def dismiss_lock_query_dialogs(page) -> None:
    remark = page.locator("#remark")
    if remark.count() and remark.is_visible():
        dialog = remark.locator(
            "xpath=ancestor::div[contains(concat(' ', normalize-space(@class), ' '), ' ui-dialog ')][1]"
        )
        cancel = dialog.get_by_role("button", name="取消", exact=True)
        if cancel.is_visible():
            cancel.click()
    for _ in range(3):
        dialogs = page.locator(".ui-dialog:visible")
        if dialogs.count() == 0:
            break
        dialog = dialogs.nth(dialogs.count() - 1)
        confirm = dialog.get_by_role("button", name="确定", exact=True)
        cancel = dialog.get_by_role("button", name="取消", exact=True)
        if confirm.count() and confirm.is_visible():
            confirm.click()
        elif cancel.count() and cancel.is_visible():
            cancel.click()
        else:
            break


def recover_conflicting_lock(page, record: dict, before_unlock=None) -> tuple:
    candidate = None
    try:
        rows = open_lock_query_from_conflict(page, record)
        candidate, error = choose_unlock_candidate(rows, record)
        if error:
            try:
                return_to_import_from_query(page)
            except Exception:
                pass
            return False, candidate, error
        if before_unlock:
            before_unlock(candidate)
        unlock_message = unlock_query_candidate(page, candidate, record)
        remaining_rows = run_locked_record_query(page, record["员工号"])
        if any(same_query_record(row, candidate) for row in remaining_rows):
            return False, candidate, "门户提示成功，但旧记录仍处于已锁查询结果"
        return_to_import_from_query(page)
        return True, candidate, unlock_message
    except Exception as error:
        try:
            dismiss_lock_query_dialogs(page)
            return_to_import_from_query(page)
        except Exception:
            pass
        return False, candidate, str(error)


def raw_table_rows(page, selector: str) -> list:
    rows = []
    row_locators = page.locator(selector)
    for row_index in range(row_locators.count()):
        cells = row_locators.nth(row_index).locator("th, td")
        values = []
        for cell_index in range(cells.count()):
            values.append(normalize_text(cells.nth(cell_index).inner_text(timeout=1000)))
        if any(values):
            rows.append(values)
    return rows


def read_quota_for_type(page, leave_type: str, year: int) -> int:
    quota_type_text = {
        "RECU_LVE": "健康疗养",
        "ALV_FD": "年假（公休假）",
    }[leave_type]
    select_leave_type(page, leave_type)
    page.locator("#nonproductionTaskRulesTipsDiv").wait_for(state="visible", timeout=15000)
    page.wait_for_function(
        """expectedText => {
            const body = document.querySelector('#nonproductionTaskRulesTipsDiv .bDiv table');
            return body && body.innerText.includes(expectedText);
        }""",
        arg=quota_type_text,
        timeout=15000,
    )
    headers = table_headers(page, "#nonproductionTaskRulesTipsDiv .hDiv table th")
    values = raw_table_rows(page, "#nonproductionTaskRulesTipsDiv .bDiv table tr")
    rows = parse_quota_rows(headers, values)
    return available_days_for_year(rows, year)


def read_page_lock_days(page, expected_days: int) -> int:
    page.wait_for_function(
        """days => document.querySelector('#lockDays1')?.value === String(days)""",
        arg=expected_days,
        timeout=10000,
    )
    value = normalize_text(page.locator("#lockDays1").input_value())
    if not re.fullmatch(r"\d+", value):
        raise RuntimeError(f"页面锁班天数异常 [{value}]")
    return int(value)


def preflight_route(page, record: dict) -> tuple:
    problem = validate_record(record)
    if problem:
        return [], {}, problem

    clear_form(page)
    fill_employee(page, record["员工号"], record.get("姓名"))
    primary_type = record["请假类型"]
    quotas = {}
    if primary_type in SMART_LEAVE_TYPES:
        year = parse_iso_date(record["开始日期"]).year
        alternate_type = ALTERNATE_LEAVE_TYPE[primary_type]
        for leave_type in (primary_type, alternate_type):
            quotas[leave_type] = read_quota_for_type(page, leave_type, year)

    start = parse_iso_date(record["开始日期"])
    end = parse_iso_date(record["结束日期"])
    expected_days = (end - start).days + 1
    select_leave_type(page, primary_type)
    fill_dates(page, record["开始日期"], record["结束日期"])
    page_days = read_page_lock_days(page, expected_days)
    if page_days != expected_days:
        return [], quotas, f"页面锁班天数{page_days}与日期计算{expected_days}不一致"

    segments, route_error = route_record(record, quotas)
    return segments, quotas, route_error


def no_related_info(row: dict) -> bool:
    return "没有相关信息" in row.get("_text", "")


def result_row_matches_record(row: dict, record: dict) -> bool:
    if no_related_info(row):
        return False

    row_text = row.get("_text", "")
    employee_id = record.get("员工号", "")
    name = record.get("姓名", "")
    start_date = record.get("开始日期", "")
    end_date = record.get("结束日期", "")
    expected_type = record.get("请假类型", "")

    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")

    if result_employee_id:
        if result_employee_id != employee_id:
            return False
    elif employee_id not in row_text:
        return False

    if name:
        if result_name:
            if not names_match(name, result_name):
                return False
        elif name not in row_text:
            return False

    if not leave_types_match(expected_type, result_type):
        return False
    if not result_type and leave_type_name(expected_type) not in row_text:
        return False

    if result_start:
        if not same_day(result_start, start_date):
            return False
    elif start_date not in row_text:
        return False

    if result_end:
        if not same_day(result_end, end_date):
            return False
    elif end_date not in row_text:
        return False

    return True


def first_matching_row(rows: list, record: dict) -> dict | None:
    for row in rows:
        if result_row_matches_record(row, record):
            return row
    return None


def read_submit_result(page, record: dict) -> tuple:
    result_headers = table_headers(page, "#showNonproductionTaskImportResultPage1 th")
    conflict_headers = table_headers(page, "#showNonproductionTaskImportResultPage2 th")
    result_rows = table_rows(page, "#showNonproductionTaskImportResultPage1 tbody.list tr", result_headers)
    conflict_rows = table_rows(page, "#showNonproductionTaskImportResultPage2 tbody.list tr", conflict_headers)
    real_results = [row for row in result_rows if not no_related_info(row)]
    real_conflicts = [row for row in conflict_rows if not no_related_info(row)]
    matching_result = first_matching_row(real_results, record)
    matching_conflict = first_matching_row(real_conflicts, record)

    if real_conflicts:
        if not matching_conflict:
            first_text = real_conflicts[0].get("_text", "") if real_conflicts else ""
            return "异常", {}, f"冲突结果未匹配当前人员: {record.get('员工号', '')} {record.get('姓名', '')}; 首行{first_text[:120]}"
        conflict = matching_conflict.get("冲突") or matching_conflict.get("_text", "")
        return "冲突", matching_conflict, conflict
    if real_results and any(no_related_info(row) for row in conflict_rows):
        if not matching_result:
            first_text = real_results[0].get("_text", "") if real_results else ""
            return "异常", {}, f"提交结果未匹配当前人员: {record.get('员工号', '')} {record.get('姓名', '')}; 首行{first_text[:120]}"
        return "成功", matching_result, ""
    if not real_results:
        return "失败", {}, "查询结果为空"
    if not matching_result:
        first_text = real_results[0].get("_text", "") if real_results else ""
        return "异常", {}, f"提交结果未匹配当前人员: {record.get('员工号', '')} {record.get('姓名', '')}; 首行{first_text[:120]}"
    note = "; ".join(row.get("_text", "") for row in conflict_rows if row.get("_text")) or "未知结果"
    return "失败", matching_result, note


def submit_and_read_result(page, record: dict) -> tuple:
    page.get_by_role("button", name="下一步").wait_for()
    page.get_by_role("button", name="下一步").click()
    page.get_by_role("button", name="继续录入").wait_for()
    status, row, remark = read_submit_result(page, record)
    if status == "成功":
        page.get_by_role("button", name="继续录入").click()
        page.locator("#showIdshowNonproductionTaskImportPage").wait_for()
    return status, row, remark


def submit_and_check(page):
    """提交并检查冲突,返回(成功, 冲突信息)"""
    # 点击下一步
    page.get_by_role("button", name="下一步").wait_for()
    page.get_by_role("button", name="下一步").click()
    # 等待继续录入按钮出现,说明页面加载完成
    page.get_by_role("button", name="继续录入").wait_for()
    # 检查查询结果是否有数据
    result_rows = page.locator("#showNonproductionTaskImportResultPage1 tbody.list tr")
    # 检查冲突列表的内容
    conflict_rows = page.locator("#showNonproductionTaskImportResultPage2 tbody.list tr")
    # 获取冲突列表的文本内容
    conflict_text = ""
    if conflict_rows.count() > 0:
        conflict_text = conflict_rows.first.inner_text()
    # 成功条件: 查询结果有数据 且 冲突列表显示"没有相关信息"
    if result_rows.count() > 0 and "没有相关信息" in conflict_text:
        # 没有冲突,点击继续录入
        page.get_by_role("button", name="继续录入").click()
        # 等待表单页面加载
        page.locator("#showIdshowNonproductionTaskImportPage").wait_for()
        return True, None
    else:
        # 有冲突或查询结果为空
        if result_rows.count() == 0:
            conflict_info = "查询结果为空"
        else:
            conflict_info = conflict_text
        return False, conflict_info


def whitelist_status(whitelist):
    """返回白名单状态文字"""
    if whitelist:
        return c_ok(f"白名单:{len(whitelist)}人")
    return c_warn("白名单:无")


def format_reason_preview(reason_text: str, limit: int = 18) -> str:
    """格式化备注预览文字"""
    preview = reason_text.replace('\n', ' / ')
    if len(preview) > limit:
        return preview[:limit] + "..."
    return preview


def reason_status(reason_text):
    """返回统一备注状态文字"""
    if reason_text:
        return c_ok(f"备注:{format_reason_preview(reason_text)}")
    return c_warn("备注:无")


def read_multiline(prompt, confirm_key='ok', cancel_key='c'):
    """读取多行输入,输入confirm_key确认,cancel_key取消"""
    print(prompt)
    lines = []
    while True:
        line = input()
        if line.lower() == cancel_key:
            return None
        if line.lower() == confirm_key:
            break
        if line:
            lines.append(line)
    if not lines:
        return None
    return '\n'.join(lines)


def set_whitelist():
    """设置白名单"""
    text = read_multiline(c_hint("请粘贴员工号列表(输入ok确认,c取消):"), 'ok', 'c')
    if text is None:
        print(c_warn("已取消"))
        return None
    wl = parse_whitelist(text)
    if not wl:
        print(c_err("未识别到有效员工号"))
        return None
    print(c_ok(f"已设置白名单,共{len(wl)}人"))
    return wl


def set_common_reason():
    """设置统一备注"""
    text = read_multiline(c_hint("请粘贴统一备注(输入OK确认,c取消):"), 'ok', 'c')
    if text is None:
        print(c_warn("本次不填写备注"))
        return None
    print(c_ok(f"已设置统一备注: {format_reason_preview(text, limit=30)}"))
    return text


def format_record(r):
    """格式化记录显示"""
    name = r['姓名'] or '未知'
    return f"{r['员工号']} {name} {r['请假类型']} {r['开始日期']}~{r['结束日期']}"


def go_back_to_form(page):
    """从提交结果页或锁班查询页返回录入表单。"""
    if page.locator("#showIdshowNonproductionTaskImportPage").is_visible():
        return
    try:
        continue_button = page.get_by_role("button", name="继续录入", exact=True)
        if continue_button.is_visible():
            continue_button.click()
        else:
            return_to_import_from_query(page)
        page.locator("#showIdshowNonproductionTaskImportPage").wait_for()
    except Exception:
        pass


def print_failed_records(failed_records):
    """打印失败记录并写入日志文件"""
    if failed_records:
        print(c_err(f"本次失败{len(failed_records)}条:"))
        for r, reason in failed_records:
            print(c_err(f"  {format_record(r)} - {reason}"))
        # 写入日志文件
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"failed_{timestamp}.txt"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(f"失败记录 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"共{len(failed_records)}条\n")
                f.write("-" * 50 + "\n")
                for r, reason in failed_records:
                    f.write(f"{format_record(r)} - {reason}\n")
            filepath = os.path.abspath(filename)
            print(c_warn(f"失败记录已保存: {filepath}"))
        except Exception as e:
            print(c_warn(f"保存日志失败: {e}"))


def create_result_excel(label: str) -> str | None:
    if not HAS_OPENPYXL:
        print(c_warn("未安装openpyxl，跳过结果Excel记录"))
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r'[\\/:*?"<>|]+', "_", label or "lock")
    output_file = os.path.abspath(f"{safe_label}_结果_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_HEADERS)
    wb.save(output_file)
    wb.close()
    print(c_ok(f"结果Excel已创建: {output_file}"))
    return output_file


def append_result_excel(
    output_file: str | None,
    original_sequence: int,
    segment_sequence: int,
    record: dict,
    segment: dict,
    quotas: dict,
    status: str,
    row: dict = None,
    remark: str = "",
    attempt: int = 1,
    recovery: str = "",
    unlocked_row: dict = None,
    excel_note: str = "",
):
    if not output_file:
        return
    row = row or {}
    segment = segment or {}
    unlocked_row = unlocked_row or {}
    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")
    input_type = record.get("请假类型", "")
    actual_type = segment.get("请假类型", "")
    alternate_type = ALTERNATE_LEAVE_TYPE.get(input_type, "")
    has_result = bool(row)
    employee_match = result_employee_id == record.get("员工号", "") if has_result else None
    name_match = names_match(record.get("姓名", ""), result_name) if has_result else None
    date_match = (
        same_day(result_start, segment.get("开始日期", ""))
        and same_day(result_end, segment.get("结束日期", ""))
    ) if has_result else None
    type_match = leave_types_match(actual_type, result_type) if has_result else None
    conflict = remark if status == "冲突" else ""
    note = excel_note or ("" if status == "成功" else remark)

    def match_text(value):
        if value is None:
            return ""
        return "是" if value else "否"

    wb = load_workbook(output_file)
    ws = wb.active
    ws.append([
        original_sequence,
        segment_sequence,
        record.get("员工号", ""),
        record.get("姓名", ""),
        leave_type_name(input_type),
        record.get("开始日期", ""),
        record.get("结束日期", ""),
        leave_type_name(actual_type),
        segment.get("开始日期", ""),
        segment.get("结束日期", ""),
        segment.get("计划天数", 0),
        quotas.get(input_type, ""),
        quotas.get(alternate_type, ""),
        status,
        row.get("锁班结果") or row.get("锁班状态") or status,
        result_name,
        result_type,
        result_start,
        result_end,
        conflict,
        note,
        match_text(employee_match),
        match_text(name_match),
        match_text(date_match),
        match_text(type_match),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        attempt,
        recovery,
        unlocked_row.get("序号", ""),
        unlocked_row.get("状态", ""),
        unlocked_row.get("员工号", ""),
        unlocked_row.get("姓名", ""),
        unlocked_row.get("开始日期", ""),
        unlocked_row.get("结束日期", ""),
        unlocked_row.get("锁班天数", ""),
        unlocked_row.get("锁班类型", ""),
        unlocked_row.get("锁班名称", ""),
        unlocked_row.get("锁班原因", ""),
        unlocked_row.get("录入人", ""),
        unlocked_row.get("录入时间", ""),
    ])
    wb.save(output_file)
    wb.close()


def append_unexecuted_segments(output_file, original_sequence, record, segments, quotas, start_index, reason):
    for segment_index in range(start_index, len(segments)):
        append_result_excel(
            output_file,
            original_sequence,
            segment_index + 1,
            record,
            segments[segment_index],
            quotas,
            "未执行",
            {},
            reason,
        )


def print_route_plan(record: dict, quotas: dict, segments: list):
    primary_type = record["请假类型"]
    if primary_type in SMART_LEAVE_TYPES:
        alternate_type = ALTERNATE_LEAVE_TYPE[primary_type]
        print(c_info(
            f"额度: {leave_type_name(primary_type)}{quotas[primary_type]}天 | "
            f"{leave_type_name(alternate_type)}{quotas[alternate_type]}天"
        ))
    else:
        print(c_info("该类型不参与额度路由，按输入类型和日期原样录入"))
    for segment_index, segment in enumerate(segments, start=1):
        print(c_ok(
            f"  片段{segment_index}: {leave_type_name(segment['请假类型'])} "
            f"{segment['开始日期']}~{segment['结束日期']} ({segment['计划天数']}天)"
        ))


def process_smart_record(
    page,
    record,
    original_sequence,
    output_file,
    common_reason,
    conflict_recovery=False,
) -> tuple:
    try:
        segments, quotas, route_error = preflight_route(page, record)
    except Exception as error:
        reason = f"录入预检失败: {error}"
        append_result_excel(
            output_file, original_sequence, 0, record, {}, {}, "预检异常", {}, reason
        )
        return False, reason

    if route_error:
        status = "额度不足" if "合计" in route_error else "预检失败"
        append_result_excel(
            output_file, original_sequence, 0, record, {}, quotas, status, {}, route_error
        )
        return False, route_error

    print_route_plan(record, quotas, segments)
    for segment_index, segment in enumerate(segments, start=1):
        attempt = 1
        unlocked_row = {}
        recovery_note = ""
        excel_note = ""
        print(c_info(
            f"提交片段{segment_index}/{len(segments)}: "
            f"{leave_type_name(segment['请假类型'])} "
            f"{segment['开始日期']}~{segment['结束日期']}"
        ))
        try:
            fill_form(
                page,
                segment["员工号"],
                segment["请假类型"],
                segment["开始日期"],
                segment["结束日期"],
                common_reason,
                segment.get("姓名"),
            )
            read_page_lock_days(page, segment["计划天数"])
            status, result_row, remark = submit_and_read_result(page, segment)
            append_result_excel(
                output_file,
                original_sequence,
                segment_index,
                record,
                segment,
                quotas,
                status,
                result_row,
                remark,
                attempt=attempt,
            )
            if status == "冲突" and conflict_recovery:
                print(c_warn("当前片段发生冲突，正在查询唯一重叠的已锁记录"))

                def persist_candidate(candidate):
                    append_result_excel(
                        output_file,
                        original_sequence,
                        segment_index,
                        record,
                        segment,
                        quotas,
                        "准备解锁",
                        result_row,
                        remark,
                        attempt=attempt,
                        recovery="已唯一定位旧记录，解锁前已落盘",
                        unlocked_row=candidate,
                    )

                recovered, unlocked_row, recovery_note = recover_conflicting_lock(
                    page,
                    segment,
                    before_unlock=persist_candidate,
                )
                unlocked_row = unlocked_row or {}
                if not recovered:
                    reason = f"冲突回退失败: {recovery_note}"
                    append_result_excel(
                        output_file,
                        original_sequence,
                        segment_index,
                        record,
                        segment,
                        quotas,
                        "冲突回退失败",
                        result_row,
                        f"{remark}; {reason}",
                        attempt=attempt,
                        recovery=reason,
                        unlocked_row=unlocked_row,
                    )
                    go_back_to_form(page)
                    append_unexecuted_segments(
                        output_file,
                        original_sequence,
                        record,
                        segments,
                        quotas,
                        segment_index,
                        f"前序片段未成功: {reason}",
                    )
                    return False, reason

                attempt = 2
                excel_note = format_unlocked_record_excel_note(unlocked_row)
                print(c_info(f"旧记录已解锁，重提片段{segment_index}"))
                fill_form(
                    page,
                    segment["员工号"],
                    segment["请假类型"],
                    segment["开始日期"],
                    segment["结束日期"],
                    common_reason,
                    segment.get("姓名"),
                )
                read_page_lock_days(page, segment["计划天数"])
                status, result_row, remark = submit_and_read_result(page, segment)
                append_result_excel(
                    output_file,
                    original_sequence,
                    segment_index,
                    record,
                    segment,
                    quotas,
                    status,
                    result_row,
                    remark,
                    attempt=attempt,
                    recovery=f"{recovery_note}; 已重提一次",
                    unlocked_row=unlocked_row,
                    excel_note=excel_note,
                )
            if status != "成功":
                reason = (
                    f"解锁旧记录后重提仍未成功: {remark or status}"
                    if attempt == 2
                    else remark or status
                )
                go_back_to_form(page)
                append_unexecuted_segments(
                    output_file,
                    original_sequence,
                    record,
                    segments,
                    quotas,
                    segment_index,
                    f"前序片段未成功: {reason}",
                )
                return False, reason
            print(c_ok(f"片段{segment_index}提交成功"))
        except Exception as error:
            reason = f"片段{segment_index}异常: {error}"
            append_result_excel(
                output_file,
                original_sequence,
                segment_index,
                record,
                segment,
                quotas,
                "异常",
                {},
                reason,
                attempt=attempt,
                recovery=recovery_note,
                unlocked_row=unlocked_row,
                excel_note=excel_note,
            )
            go_back_to_form(page)
            append_unexecuted_segments(
                output_file,
                original_sequence,
                record,
                segments,
                quotas,
                segment_index,
                f"前序片段异常: {reason}",
            )
            return False, reason
    return True, ""


def batch_mode(page, whitelist, common_reason, conflict_recovery):
    """批量粘贴模式。"""
    while True:
        print(f"{c_info('[智能路由批量]')} {whitelist_status(whitelist)} | {reason_status(common_reason)}")
        text = read_multiline(c_hint("请粘贴数据(输入ok确认,b返回):"), 'ok', 'b')
        if text is None:
            return
        records, errors = parse_batch_input(text, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(err))
        if not records:
            print(c_err("没有可处理的记录"))
            continue
        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")
        confirm = input(c_hint("y开始填写,n重新粘贴,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        result_file = create_result_excel("智能路由锁班批量")
        failed_records = process_record_list(
            page, records, result_file, common_reason, conflict_recovery
        )
        print(c_ok("批量处理完成"))
        if result_file:
            print(c_ok(f"结果Excel: {result_file}"))
        print_failed_records(failed_records)
        return


def manual_mode(page, whitelist, common_reason, conflict_recovery):
    """手动单条模式。"""
    result_file = None
    sequence = 0
    while True:
        print(f"{c_info('[手动模式]')} {whitelist_status(whitelist)} | {reason_status(common_reason)} | {c_hint('粘贴数据,b返回主菜单:')}")
        text = input().strip()
        if text.lower() == 'b':
            return
        if not text:
            continue
        record = parse_single_record(text)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            print(c_err("该员工不在白名单中"))
            continue
        problem = validate_record(record)
        if problem:
            print(c_err(problem))
            continue
        print(f"待处理: {format_record(record)}")
        confirm = input(c_hint("y确认智能路由并提交，其他键取消: ")).strip().lower()
        if confirm != 'y':
            continue
        if result_file is None:
            result_file = create_result_excel("智能路由锁班手动")
        sequence += 1
        success, reason = process_smart_record(
            page, record, sequence, result_file, common_reason, conflict_recovery
        )
        if not success:
            beep_error()
            print(c_err(reason))


def excel_mode(page, whitelist, common_reason, conflict_recovery):
    """Excel导入模式"""
    if not HAS_OPENPYXL:
        print(c_err("未安装openpyxl库，请运行: pip install openpyxl"))
        return

    failed_records = []
    while True:
        print(f"{c_info('[Excel导入]')} {whitelist_status(whitelist)} | {reason_status(common_reason)}")
        filepath = input(c_hint("请输入Excel文件路径(b返回): ")).strip()
        if filepath.lower() == 'b':
            print_failed_records(failed_records)
            return

        # 去除引号
        filepath = filepath.strip('"').strip("'")

        if not os.path.exists(filepath):
            print(c_err("文件不存在"))
            continue

        records, errors = parse_excel_file(filepath, whitelist)
        if errors:
            print(c_err("解析错误:"))
            for err in errors:
                print(c_err(f"  {err}"))

        if not records:
            print(c_err("没有可处理的记录"))
            continue

        print(c_ok(f"共{len(records)}条有效数据:"))
        for i, r in enumerate(records, 1):
            print(f"{i}. {format_record(r)}")

        confirm = input(c_hint("y开始填写,n重新选择,b返回主菜单: ")).strip().lower()
        if confirm == 'b':
            return
        if confirm != 'y':
            continue
        result_file = create_result_excel("智能路由锁班Excel")
        failed_records = process_record_list(
            page, records, result_file, common_reason, conflict_recovery
        )
        print(c_ok("Excel导入完成"))
        if result_file:
            print(c_ok(f"结果Excel: {result_file}"))
        print_failed_records(failed_records)
        return


def process_record_list(page, records, result_file, common_reason, conflict_recovery=False):
    failed_records = []
    for sequence, record in enumerate(records, start=1):
        print(f"{c_info(f'[{sequence}/{len(records)}]')} 预检: {format_record(record)}")
        success, reason = process_smart_record(
            page,
            record,
            sequence,
            result_file,
            common_reason,
            conflict_recovery,
        )
        if not success:
            beep_error()
            print(c_err(reason))
            failed_records.append((record, reason))
    return failed_records


def main():
    print(c_info("锁班皇帝 - 智能路由助手"))
    print(c_info("支持全部锁班类型；健康疗养与飞行员公休（订座）按可休天数自动分配"))
    if not HAS_OPENPYXL:
        print(c_err("缺少openpyxl，无法生成实时结果文件。请先运行: pip install -r requirements.txt"))
        return
    # 浏览器路径
    browser_path = input(c_hint("浏览器路径(回车用默认): ")).strip() or None
    if browser_path:
        print(c_ok(f"使用指定浏览器: {browser_path}"))
    else:
        print(c_ok("使用默认浏览器"))
    # 白名单
    whitelist = None
    use_wl = input(c_hint("是否预设白名单?(y/n): ")).strip().lower()
    if use_wl == 'y':
        whitelist = set_whitelist()
    else:
        print(c_ok("不设置白名单,处理所有员工"))
    common_reason = None
    use_reason = input(c_hint("是否填写统一备注?(y/n): ")).strip().lower()
    if use_reason == 'y':
        common_reason = set_common_reason()
    else:
        print(c_ok("本次不填写备注"))
    recovery_answer = input(c_hint(
        "是否启用冲突自动解锁并重提? 仅唯一匹配已锁记录时执行(y/n): "
    )).strip().lower()
    conflict_recovery = recovery_answer == 'y'
    if conflict_recovery:
        print(c_warn("已启用冲突回退：命中唯一已锁记录时会解锁旧整行并重提一次"))
    else:
        print(c_ok("本次不自动解锁冲突记录"))
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=False, executable_path=browser_path)
    context = browser.new_context()
    context.set_default_timeout(30000)
    page = context.new_page()
    # 登录
    try:
        page.goto("https://ieb.csair.com/login")
        page.wait_for_load_state("networkidle")
        page.locator("#scanLogin").wait_for()
        page.locator("#scanLogin").click()
        print(c_info("请扫码登录..."))
        page.wait_for_url("**/index/**", timeout=120000)
        page.wait_for_load_state("networkidle")
        print(c_ok("登录成功"))
    except Exception as e:
        print(c_err(f"自动登录失败: {e}"))
        print(c_warn("请手动完成登录"))
        input(c_hint("登录完成后按回车继续..."))
    # 导航到非生产任务录入页面
    try:
        print(c_info("正在进入非生产任务录入页面..."))
        page.goto("https://ieb.csair.com/index/index")
        page.wait_for_load_state("networkidle")
        page.get_by_text("运行管理").nth(1).wait_for()
        page.get_by_text("运行管理").nth(1).click()
        page.get_by_role("link", name="非生产任务").wait_for()
        page.get_by_role("link", name="非生产任务").click()
        page.get_by_role("link", name="非生产任务录入").wait_for()
        page.get_by_role("link", name="非生产任务录入").click()
        page.locator("#mainContent").wait_for()
        page.locator("#mainContent").click()
        page.wait_for_load_state("networkidle")
        print(c_ok("已进入非生产任务录入页面"))
    except Exception as e:
        print(c_err(f"自动导航失败: {e}"))
        print(c_warn("请手动进入非生产任务录入页面"))
        input(c_hint("准备好后按回车继续..."))
    print(c_ok("开始工作"))
    while True:
        print(f"{whitelist_status(whitelist)} | {reason_status(common_reason)} | {c_hint('1批量 2手动 3Excel导入 w设白名单 c清白名单 q退出')}")
        cmd = input(c_hint("选择: ")).strip().lower()
        if cmd == '1':
            batch_mode(page, whitelist, common_reason, conflict_recovery)
        elif cmd == '2':
            manual_mode(page, whitelist, common_reason, conflict_recovery)
        elif cmd == '3':
            excel_mode(page, whitelist, common_reason, conflict_recovery)
        elif cmd == 'w':
            new_wl = set_whitelist()
            if new_wl is not None:
                whitelist = new_wl
        elif cmd == 'c':
            whitelist = None
            print(c_ok("已清除白名单"))
        elif cmd == 'q':
            break
    browser.close()
    pw.stop()
    print(c_info("结束"))


if __name__ == "__main__":
    main()
