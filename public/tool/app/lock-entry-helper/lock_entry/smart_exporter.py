"""Smart result workbook and unexecuted-segment evidence."""

import os
from datetime import datetime

from .common import *
try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def create_result_excel(label: str) -> str | None:
    if not HAS_OPENPYXL:
        print(c_warn("未安装openpyxl，跳过结果Excel记录"))
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r'[\\/:*?"<>|]+', "_", label or "lock")
    output_file = os.path.abspath(f"{safe_label}_结果_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_HEADERS)
    wb.save(output_file)
    wb.close()
    print(c_ok(f"结果Excel已创建: {output_file}"))
    return output_file


def append_result_excel(
    output_file: str | None,
    original_sequence: int,
    segment_sequence: int,
    record: dict,
    segment: dict,
    quotas: dict,
    status: str,
    row: dict = None,
    remark: str = "",
    attempt: int = 1,
    recovery: str = "",
    unlocked_row: dict = None,
    excel_note: str = "",
):
    if not output_file:
        return
    row = row or {}
    segment = segment or {}
    unlocked_row = unlocked_row or {}
    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")
    input_type = record.get("请假类型", "")
    actual_type = segment.get("请假类型", "")
    alternate_type = ALTERNATE_LEAVE_TYPE.get(input_type, "")
    has_result = bool(row)
    employee_match = result_employee_id == record.get("员工号", "") if has_result else None
    name_match = names_match(record.get("姓名", ""), result_name) if has_result else None
    date_match = (
        same_day(result_start, segment.get("开始日期", ""))
        and same_day(result_end, segment.get("结束日期", ""))
    ) if has_result else None
    type_match = leave_types_match(actual_type, result_type) if has_result else None
    conflict = remark if status == "冲突" else ""
    note = excel_note or ("" if status == "成功" else remark)

    def match_text(value):
        if value is None:
            return ""
        return "是" if value else "否"

    wb = load_workbook(output_file)
    ws = wb.active
    ws.append([
        original_sequence,
        segment_sequence,
        record.get("员工号", ""),
        record.get("姓名", ""),
        leave_type_name(input_type),
        record.get("开始日期", ""),
        record.get("结束日期", ""),
        leave_type_name(actual_type),
        segment.get("开始日期", ""),
        segment.get("结束日期", ""),
        segment.get("计划天数", 0),
        quotas.get(input_type, ""),
        quotas.get(alternate_type, ""),
        status,
        row.get("锁班结果") or row.get("锁班状态") or status,
        result_name,
        result_type,
        result_start,
        result_end,
        conflict,
        note,
        match_text(employee_match),
        match_text(name_match),
        match_text(date_match),
        match_text(type_match),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        attempt,
        recovery,
        unlocked_row.get("序号", ""),
        unlocked_row.get("状态", ""),
        unlocked_row.get("员工号", ""),
        unlocked_row.get("姓名", ""),
        unlocked_row.get("开始日期", ""),
        unlocked_row.get("结束日期", ""),
        unlocked_row.get("锁班天数", ""),
        unlocked_row.get("锁班类型", ""),
        unlocked_row.get("锁班名称", ""),
        unlocked_row.get("锁班原因", ""),
        unlocked_row.get("录入人", ""),
        unlocked_row.get("录入时间", ""),
    ])
    wb.save(output_file)
    wb.close()


def append_unexecuted_segments(output_file, original_sequence, record, segments, quotas, start_index, reason):
    for segment_index in range(start_index, len(segments)):
        append_result_excel(
            output_file,
            original_sequence,
            segment_index + 1,
            record,
            segments[segment_index],
            quotas,
            "未执行",
            {},
            reason,
        )
