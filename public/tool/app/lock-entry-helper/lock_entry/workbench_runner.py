"""Persistent browser worker for the original and smart lock-entry workbenches."""

from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .common import leave_type_name
from .exporter import append_result_excel as append_original_result
from .exporter import create_result_excel as create_original_result
from .launcher import start_portal_session
from .original_runner import go_back_to_form as original_go_back_to_form
from .portal import fill_form as fill_original_form
from .portal import submit_and_read_result as submit_original_result
from .smart_exporter import create_result_excel as create_smart_result
from .smart_runner import process_smart_record
from .workbench_input import RunConfig, read_input


@dataclass(frozen=True)
class RunSummary:
    total: int
    success: int
    failed: int


def _browser_closed(error: Exception) -> bool:
    message = str(error).lower()
    return any(
        marker in message
        for marker in (
            "target page, context or browser has been closed",
            "browser has been closed",
            "page has been closed",
        )
    )


def _progress_event(
    total: int,
    completed: int,
    success: int,
    failed: int,
    current: str = "",
) -> dict[str, Any]:
    return {
        "type": "progress",
        "total": total,
        "completed": completed,
        "success": success,
        "failed": failed,
        "current": current,
    }


def _record_event(
    index: int,
    record: dict,
    status: str,
    result_row: dict | None = None,
    remark: str = "",
) -> dict[str, Any]:
    row = result_row or {}
    return {
        "type": "record_result",
        "index": index,
        "segmentIndex": 1,
        "employeeId": record.get("员工号", ""),
        "name": record.get("姓名", "") or "",
        "inputType": leave_type_name(record.get("请假类型", "")),
        "actualType": row.get("锁班类型", "") or leave_type_name(record.get("请假类型", "")),
        "inputStartDate": record.get("开始日期", ""),
        "inputEndDate": record.get("结束日期", ""),
        "actualStartDate": row.get("开始日期", "") or record.get("开始日期", ""),
        "actualEndDate": row.get("结束日期", "") or record.get("结束日期", ""),
        "status": status,
        "portalStatus": row.get("锁班结果") or row.get("锁班状态") or status,
        "remark": remark,
        "attempt": 1,
        "recovery": "",
    }


def run_original_records(
    records: list[dict],
    page: Any,
    output_file: str,
    common_reason: str,
    emit: Callable[[dict[str, Any]], None],
) -> RunSummary:
    success = 0
    failed = 0
    total = len(records)
    for index, record in enumerate(records, start=1):
        current = f"{record.get('员工号', '')} {record.get('姓名', '') or ''}".strip()
        emit(_progress_event(total, index - 1, success, failed, current))
        emit({"type": "log", "level": "info", "message": f"[{index}/{total}] 正在录入 {current}"})
        try:
            fill_original_form(
                page,
                record["员工号"],
                record["请假类型"],
                record["开始日期"],
                record["结束日期"],
                common_reason or None,
            )
            status, result_row, remark = submit_original_result(page, record)
            append_original_result(output_file, index, record, status, result_row, remark)
            emit(_record_event(index, record, status, result_row, remark))
            if status == "成功":
                success += 1
                emit({"type": "log", "level": "success", "message": f"{current} 提交成功"})
            else:
                failed += 1
                original_go_back_to_form(page)
                emit(
                    {
                        "type": "log",
                        "level": "error",
                        "message": f"{current} {status}：{remark or '门户未返回原因'}",
                    }
                )
        except Exception as error:
            failed += 1
            append_original_result(output_file, index, record, "异常", {}, str(error))
            emit(_record_event(index, record, "异常", {}, str(error)))
            emit({"type": "log", "level": "error", "message": f"{current} 录入异常：{error}"})
            if _browser_closed(error):
                raise
        emit(_progress_event(total, index, success, failed))
    return RunSummary(total=total, success=success, failed=failed)


def _workbook_rows_after(output_file: str, row_number: int) -> list[dict]:
    workbook = load_workbook(output_file, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        headers = [str(cell.value or "") for cell in sheet[1]]
        return [
            dict(zip(headers, values))
            for values in sheet.iter_rows(min_row=row_number + 1, values_only=True)
        ]
    finally:
        workbook.close()


def _workbook_max_row(output_file: str) -> int:
    workbook = load_workbook(output_file, data_only=True, read_only=True)
    try:
        return workbook.active.max_row
    finally:
        workbook.close()


def _smart_record_event(row: dict) -> dict[str, Any]:
    return {
        "type": "record_result",
        "index": row.get("原始序号", 0),
        "segmentIndex": row.get("片段序号", 0),
        "employeeId": row.get("员工号", ""),
        "name": row.get("姓名", "") or "",
        "inputType": row.get("输入锁班类型", "") or "",
        "actualType": row.get("实际锁班类型", "") or "",
        "inputStartDate": row.get("输入开始日期", "") or "",
        "inputEndDate": row.get("输入结束日期", "") or "",
        "actualStartDate": row.get("实际开始日期", "") or "",
        "actualEndDate": row.get("实际结束日期", "") or "",
        "status": row.get("处理状态", "") or "",
        "portalStatus": row.get("锁班结果", "") or "",
        "remark": row.get("备注", "") or row.get("冲突", "") or "",
        "attempt": row.get("尝试次数", 1) or 1,
        "recovery": row.get("冲突回退", "") or "",
    }


def run_smart_records(
    records: list[dict],
    page: Any,
    output_file: str,
    common_reason: str,
    conflict_recovery: bool,
    emit: Callable[[dict[str, Any]], None],
) -> RunSummary:
    success = 0
    failed = 0
    total = len(records)
    for index, record in enumerate(records, start=1):
        current = f"{record.get('员工号', '')} {record.get('姓名', '') or ''}".strip()
        emit(_progress_event(total, index - 1, success, failed, current))
        emit({"type": "log", "level": "info", "message": f"[{index}/{total}] 正在预检并录入 {current}"})
        previous_row = _workbook_max_row(output_file)
        completed, reason = process_smart_record(
            page,
            record,
            index,
            output_file,
            common_reason or None,
            conflict_recovery,
        )
        for row in _workbook_rows_after(output_file, previous_row):
            emit(_smart_record_event(row))
        if completed:
            success += 1
            emit({"type": "log", "level": "success", "message": f"{current} 全部片段提交成功"})
        else:
            failed += 1
            emit({"type": "log", "level": "error", "message": f"{current} 未完成：{reason}"})
        emit(_progress_event(total, index, success, failed))
    return RunSummary(total=total, success=success, failed=failed)


def run_batch(
    config: RunConfig,
    page: Any,
    emit: Callable[[dict[str, Any]], None],
) -> None:
    try:
        records, input_errors = read_input(config)
        for message in input_errors:
            emit({"type": "log", "level": "error", "message": message})
        if not records:
            raise ValueError("没有可处理的记录")

        label = "智能路由锁班工作台" if config.mode == "smart" else "原始锁班工作台"
        create_result = create_smart_result if config.mode == "smart" else create_original_result
        output_file = create_result(label, config.output_directory)
        if not output_file:
            raise RuntimeError("无法创建结果 Excel，请确认已安装 openpyxl")
        emit({"type": "result", "path": output_file})
        emit({"type": "status", "phase": "running", "message": f"正在串行录入 {len(records)} 条有效数据"})

        if config.mode == "smart":
            summary = run_smart_records(
                records,
                page,
                output_file,
                config.common_reason,
                config.conflict_recovery,
                emit,
            )
        else:
            summary = run_original_records(
                records,
                page,
                output_file,
                config.common_reason,
                emit,
            )
        emit(
            {
                "type": "completed",
                "message": (
                    f"录入完成：成功 {summary.success}，失败 {summary.failed}；"
                    "浏览器保持打开，可重新准备下一批"
                ),
                "total": summary.total,
                "success": summary.success,
                "failed": summary.failed,
                "path": output_file,
            }
        )
    except Exception as error:
        if _browser_closed(error):
            raise
        emit({"type": "failed", "message": str(error)})


def run_worker(config: RunConfig, event_queue: Any, command_queue: Any | None = None) -> None:
    emit = event_queue.put
    session = None
    try:
        emit(
            {
                "type": "log",
                "level": "info",
                "message": "当前模式：智能串行" if config.mode == "smart" else "当前模式：原始串行",
            }
        )
        session = start_portal_session(
            config.browser_path or None,
            default_timeout=30000,
            login_timeout=600000,
            emit=emit,
            interactive_fallback=False,
        )
        current_config = config
        if command_queue is None:
            run_batch(current_config, session.page, emit)
            return
        if current_config.auto_run:
            run_batch(current_config, session.page, emit)

        while True:
            command = command_queue.get()
            command_name = command.get("command", "") if isinstance(command, dict) else str(command)
            next_config = command.get("config") if isinstance(command, dict) else None
            if command_name == "prepare" and isinstance(next_config, RunConfig):
                if next_config.mode != current_config.mode:
                    emit({"type": "failed", "message": "工作台模式与启动入口不一致"})
                    continue
                current_config = next_config
                emit(
                    {
                        "type": "status",
                        "phase": "prepared",
                        "message": "新一批数据已准备，浏览器保持打开，等待开始录入",
                    }
                )
                if current_config.auto_run:
                    run_batch(current_config, session.page, emit)
            elif command_name == "run":
                run_batch(current_config, session.page, emit)
    except Exception as error:
        emit({"type": "failed", "message": str(error)})
    finally:
        if session is not None:
            session.close()


def run_worker_process(config: RunConfig, event_queue: Any, command_queue: Any | None = None) -> None:
    if os.name != "nt":
        try:
            os.setsid()
        except OSError:
            pass
    run_worker(config, event_queue, command_queue)
