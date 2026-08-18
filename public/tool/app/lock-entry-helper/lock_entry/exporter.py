"""Original result workbook persistence."""

import os
import re
from datetime import datetime

from .common import c_err, c_ok, c_warn, leave_type_name, leave_types_match, names_match, same_day
try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ORIGINAL_RESULT_HEADERS = [
    "序号",
    "员工号",
    "姓名",
    "锁班类型",
    "开始日期",
    "结束日期",
    "处理状态",
    "锁班结果",
    "结果姓名",
    "结果锁班类型",
    "结果开始日期",
    "结果结束日期",
    "冲突",
    "备注",
    "员工号匹配",
    "姓名匹配",
    "日期匹配",
    "类型匹配",
    "处理时间",
]
RESULT_HEADERS = ORIGINAL_RESULT_HEADERS

def create_result_excel(label: str) -> str | None:
    if not HAS_OPENPYXL:
        print(c_warn("未安装openpyxl，跳过结果Excel记录"))
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r'[\\/:*?"<>|]+', "_", label or "lock")
    output_file = os.path.abspath(f"{safe_label}_结果_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_HEADERS)
    wb.save(output_file)
    wb.close()
    print(c_ok(f"结果Excel已创建: {output_file}"))
    return output_file


def append_result_excel(output_file: str | None, sequence: int, record: dict, status: str, row: dict = None, remark: str = ""):
    if not output_file:
        return
    row = row or {}
    result_employee_id = row.get("员工号", "")
    result_name = row.get("姓名", "")
    result_start = row.get("开始日期", "")
    result_end = row.get("结束日期", "")
    result_type = row.get("锁班类型", "")
    expected_type_name = leave_type_name(record.get("请假类型", ""))
    employee_match = (not result_employee_id) or result_employee_id == record.get("员工号", "")
    name_match = names_match(record.get("姓名", ""), result_name)
    date_match = (not result_start or same_day(result_start, record.get("开始日期", ""))) and (
        not result_end or same_day(result_end, record.get("结束日期", ""))
    )
    type_match = leave_types_match(record.get("请假类型", ""), result_type)
    conflict = remark if status == "冲突" else ""
    note = "" if status == "成功" else remark

    wb = load_workbook(output_file)
    ws = wb.active
    ws.append([
        sequence,
        record.get("员工号", ""),
        record.get("姓名", ""),
        expected_type_name,
        record.get("开始日期", ""),
        record.get("结束日期", ""),
        status,
        row.get("锁班结果") or row.get("锁班状态") or status,
        result_name,
        result_type,
        result_start,
        result_end,
        conflict,
        note,
        "是" if employee_match else "否",
        "是" if name_match else "否",
        "是" if date_match else "否",
        "是" if type_match else "否",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(output_file)
    wb.close()
