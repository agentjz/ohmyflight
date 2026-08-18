"""Smart helper Excel header and business-date parsing."""

import re
from datetime import date, datetime, timedelta

from .common import normalize_date, normalize_text, parse_leave_type
from .constants import EXCEL_HEADER_ALIASES
try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def normalize_employee_id(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_portal_business_date(value: str) -> date:
    """从门户日期或日期时间文本读取业务日期。"""
    text = normalize_text(value)
    match = re.match(r"^(\d{4}-\d{2}-\d{2})(?:\s|$)", text)
    if not match:
        raise ValueError(f"门户日期格式异常 [{text}]")
    return datetime.strptime(match.group(1), "%Y-%m-%d").date()


def date_ranges_overlap(
    first_start: str,
    first_end: str,
    second_start: str,
    second_end: str,
) -> bool:
    """按包含首尾的自然日区间判断两条锁班是否重叠。"""
    first_start_date = parse_portal_business_date(first_start)
    first_end_date = parse_portal_business_date(first_end)
    second_start_date = parse_portal_business_date(second_start)
    second_end_date = parse_portal_business_date(second_end)
    if first_end_date < first_start_date or second_end_date < second_start_date:
        raise ValueError("锁班结束日期早于开始日期")
    return first_start_date <= second_end_date and second_start_date <= first_end_date


def choose_unlock_candidate(rows: list, record: dict) -> tuple:
    """只接受员工号、已锁状态和日期区间共同唯一命中的旧记录。"""
    employee_id = normalize_employee_id(record.get("员工号"))
    try:
        parse_portal_business_date(record.get("开始日期", ""))
        parse_portal_business_date(record.get("结束日期", ""))
    except ValueError as error:
        return None, f"冲突记录日期异常: {error}"

    candidates = []
    for row in rows:
        if normalize_employee_id(row.get("员工号")) != employee_id:
            continue
        if normalize_text(row.get("状态")) != "已锁":
            continue
        try:
            overlaps = date_ranges_overlap(
                row.get("开始日期", ""),
                row.get("结束日期", ""),
                record.get("开始日期", ""),
                record.get("结束日期", ""),
            )
        except ValueError as error:
            return None, f"已锁候选日期异常: {error}"
        if overlaps:
            candidates.append(row)

    if not candidates:
        return None, "未找到员工号一致、状态已锁且日期重叠的旧锁班"
    if len(candidates) != 1:
        return None, f"找到{len(candidates)}条日期重叠的已锁记录，无法唯一定位"
    return candidates[0], ""


def format_unlocked_record_excel_note(row: dict) -> str:
    return (
        f"已解锁：锁班名称{normalize_text(row.get('锁班名称'))}；"
        f"锁班原因{normalize_text(row.get('锁班原因'))}；"
        f"开始日期{normalize_text(row.get('开始日期'))}；"
        f"结束日期{normalize_text(row.get('结束日期'))}"
    )


def format_business_date(value) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        if isinstance(value, int) and re.fullmatch(r"(?:19|20)\d{6}", str(value)):
            return normalize_date(str(value))
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except (TypeError, ValueError, OverflowError):
            return ""
    if isinstance(value, str):
        return normalize_date(value)
    return ""


def excel_header_indexes(header_row: tuple) -> tuple[dict, list]:
    normalized = [normalize_text(value) for value in header_row]
    indexes = {}
    missing = []
    for field, aliases in EXCEL_HEADER_ALIASES.items():
        index = next((normalized.index(alias) for alias in aliases if alias in normalized), None)
        if index is None:
            missing.append(field)
        else:
            indexes[field] = index
    return indexes, missing


def parse_excel_file(filepath: str, whitelist: set = None) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    if not HAS_OPENPYXL:
        return [], ["未安装openpyxl库，请运行: pip install openpyxl"]

    records = []
    errors = []

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        indexes, missing_headers = excel_header_indexes(header_row)
        if missing_headers:
            wb.close()
            return [], [f"Excel缺少表头: {', '.join(missing_headers)}"]

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            emp_id = normalize_employee_id(row[indexes["员工号"]])
            name = normalize_text(row[indexes["姓名"]]) or None
            leave_type_raw = normalize_text(row[indexes["锁班类型"]])
            start = format_business_date(row[indexes["开始日期"]])
            end = format_business_date(row[indexes["结束日期"]]) or start

            leave_type = parse_leave_type(leave_type_raw)
            record = {
                "员工号": emp_id,
                "姓名": name,
                "请假类型": leave_type,
                "开始日期": start,
                "结束日期": end,
            }
            from .smart_router import validate_record
            problem = validate_record(record)
            if problem:
                errors.append(f"第{row_num}行: {problem}")
                continue
            if whitelist and emp_id not in whitelist:
                continue
            records.append(record)

        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")

    return records, errors
