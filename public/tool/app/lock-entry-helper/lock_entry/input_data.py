"""Original pasted, single-record, and Excel input parsing."""

import re
from datetime import datetime

from .common import normalize_date, parse_leave_type
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def parse_excel_file(filepath: str, whitelist: set = None) -> tuple:
    """解析Excel文件，返回(records, errors)"""
    if not HAS_OPENPYXL:
        return [], ["未安装openpyxl库，请运行: pip install openpyxl"]
    
    records = []
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # 跳过表头，从第2行开始
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            # 列顺序: 员工号、姓名、锁班类型、开始日期、结束日期
            emp_id = str(row[0]).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            leave_type_raw = str(row[2]).strip() if len(row) > 2 and row[2] else None
            start_date = row[3] if len(row) > 3 else None
            end_date = row[4] if len(row) > 4 else None
            
            # 验证员工号
            if not emp_id or not re.match(r'^\d{6}$', emp_id):
                if emp_id and emp_id != 'None':
                    errors.append(f"第{row_num}行: 员工号格式错误 [{emp_id}]")
                continue
            
            # 白名单过滤
            if whitelist and emp_id not in whitelist:
                continue
            
            # 解析锁班类型
            leave_type = parse_leave_type(leave_type_raw)
            if not leave_type:
                errors.append(f"第{row_num}行: 未识别锁班类型 [{leave_type_raw}]")
                continue
            
            # 解析日期
            def format_date(d):
                if isinstance(d, datetime):
                    return d.strftime('%Y-%m-%d')
                if isinstance(d, str):
                    return normalize_date(d)
                return None
            
            start = format_date(start_date)
            end = format_date(end_date) if end_date else start
            
            if not start:
                errors.append(f"第{row_num}行: 日期格式错误")
                continue
            
            records.append({
                "员工号": emp_id,
                "姓名": name,
                "请假类型": leave_type,
                "开始日期": start,
                "结束日期": end or start
            })
        
        wb.close()
    except Exception as e:
        errors.append(f"读取Excel失败: {e}")
    
    return records, errors


def parse_single_record(text: str) -> dict:
    """解析单条记录"""
    result = {"员工号": None, "姓名": None, "请假类型": None, "开始日期": None, "结束日期": None}
    emp = re.search(r'\b(\d{6})\b', text)
    if emp:
        result["员工号"] = emp.group(1)
    name = re.search(r'\d{6}\s*([\u4e00-\u9fa5]{2,4})', text)
    if name:
        result["姓名"] = name.group(1)
    result["请假类型"] = parse_leave_type(text)
    dates = re.findall(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', text)
    if dates:
        result["开始日期"] = normalize_date(dates[0])
        result["结束日期"] = normalize_date(dates[1]) if len(dates) > 1 else normalize_date(dates[0])
    return result


def split_continuous_text(text: str) -> list:
    """把连续粘贴的文本按员工号切分成多条记录"""
    # 按6位员工号切分
    parts = re.split(r'(?=\d{6}[\u4e00-\u9fa5])', text)
    return [p.strip() for p in parts if p.strip() and re.search(r'\d{6}', p)]


def parse_batch_input(text: str, whitelist: set = None) -> tuple:
    """解析批量输入"""
    records = []
    errors = []
    # 先按换行分，如果只有一行且很长，尝试按员工号切分
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
    if len(lines) == 1 and len(lines[0]) > 100:
        lines = split_continuous_text(lines[0])
    for i, line in enumerate(lines, 1):
        record = parse_single_record(line)
        if whitelist and record["员工号"] and record["员工号"] not in whitelist:
            continue
        if not record["员工号"]:
            errors.append(f"第{i}条: 未识别员工号 [{line[:50]}]")
            continue
        if not record["请假类型"]:
            errors.append(f"第{i}条: 未识别请假类型 [{line[:50]}]")
            continue
        if not record["开始日期"]:
            errors.append(f"第{i}条: 未识别日期 [{line[:50]}]")
            continue
        records.append(record)
    return records, errors
