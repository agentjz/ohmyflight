from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook

from .input_data import QueryRecord
from .portal import TableResult


INPUT_HEADERS = ["输入员工号", "输入姓名", "输入开始日期", "输入结束日期"]
STATUS_HEADERS = ["查询状态", "错误说明"]
STRIPPED_TIME_FIELDS = {"飞行时间", "飞行经历", "左座经历"}
EXCEL_DATE_FORMAT = "yyyy-mm-dd"
SCOPE_FIELDS = {
    "flight_time": "飞行时间",
    "flight_experience": "飞行经历",
    "left_seat_experience": "左座经历",
}
SCOPE_LABELS = {
    "flight_time": "飞行时间+起落数",
    "flight_experience": "飞行经历+起落数",
    "left_seat_experience": "左座经历+起落数",
    "all": "全部数据",
}
ScopeValue = str | Sequence[str]


def normalize_scope(scope: ScopeValue) -> tuple[str, ...]:
    values = [scope] if isinstance(scope, str) else list(scope)
    if not values:
        raise ValueError("至少选择一个查询范围")
    if "all" in values:
        return ("all",)
    normalized = tuple(dict.fromkeys(values))
    if any(value not in SCOPE_FIELDS for value in normalized):
        raise ValueError("不支持的查询范围")
    return normalized


def describe_scope(scope: ScopeValue) -> str:
    normalized = normalize_scope(scope)
    if normalized == ("all",):
        return SCOPE_LABELS["all"]
    return "、".join(SCOPE_LABELS[value] for value in normalized)


@dataclass(frozen=True)
class OutputPaths:
    original: Path
    stripped: Path


def strip_minutes(value):
    if not isinstance(value, str):
        return value
    match = re.fullmatch(r"\s*(\d+)\s*[:：]\s*\d{1,2}\s*", value)
    return match.group(1) if match else value


def _safe_run_name(run_id: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", run_id).strip("-")
    return safe or "run"


def _save_atomic(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _headers(worksheet) -> list[str]:
    return [str(cell.value or "").strip() for cell in worksheet[1]]


def _header_map(worksheet) -> dict[str, int]:
    return {header: index + 1 for index, header in enumerate(_headers(worksheet)) if header}


class ResultExporter:
    def __init__(self, output_directory: Path, run_id: str, scope: ScopeValue = "all"):
        output_directory.mkdir(parents=True, exist_ok=True)
        self.scope = normalize_scope(scope)
        name = _safe_run_name(run_id)
        self.paths = OutputPaths(
            original=output_directory / f"飞行经历查询_{name}_原版.xlsx",
            stripped=output_directory / f"飞行经历查询_{name}_去分钟版.xlsx",
        )

    def initialize(self, records: Iterable[QueryRecord], input_errors: list[str]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "查询结果"
        worksheet.append(INPUT_HEADERS + STATUS_HEADERS)
        for record in records:
            worksheet.append(
                [
                    record.employee_id,
                    record.name,
                    record.start_date,
                    record.end_date,
                    "待处理",
                    "",
                ]
            )
            worksheet.cell(worksheet.max_row, 3).number_format = EXCEL_DATE_FORMAT
            worksheet.cell(worksheet.max_row, 4).number_format = EXCEL_DATE_FORMAT
        if input_errors:
            error_sheet = workbook.create_sheet("输入错误")
            error_sheet.append(["来源", "错误说明"])
            for index, message in enumerate(input_errors, start=1):
                error_sheet.append([index, message])
        _save_atomic(workbook, self.paths.original)
        workbook.close()
        self._refresh_stripped()

    def _ensure_result_headers(self, worksheet, result_headers: list[str]) -> dict[str, int]:
        existing = _header_map(worksheet)
        for header in result_headers:
            if header not in existing:
                column = worksheet.max_column + 1
                worksheet.cell(1, column, header)
                existing[header] = column
        return existing

    def selected_headers(self, result: TableResult) -> list[str]:
        if self.scope == ("all",):
            return result.headers
        allowed = {"员工号", "姓名", "起落总数"}
        allowed.update(SCOPE_FIELDS[key] for key in self.scope)
        return [header for header in result.headers if header in allowed]

    def write_success(self, record_index: int, record: QueryRecord, result: TableResult) -> None:
        workbook = load_workbook(self.paths.original)
        try:
            worksheet = workbook["查询结果"]
            result_headers = self.selected_headers(result)
            missing = [] if self.scope == ("all",) else [
                SCOPE_FIELDS[key] for key in self.scope if SCOPE_FIELDS[key] not in result_headers
            ]
            if self.scope != ("all",) and "起落总数" not in result_headers:
                missing.append("起落总数")
            if missing:
                raise ValueError(f"页面缺少查询字段：{'、'.join(missing)}")
            columns = self._ensure_result_headers(worksheet, result_headers)
            row_number = record_index + 2
            for header in result_headers:
                worksheet.cell(row_number, columns[header], result.values.get(header, ""))
            worksheet.cell(row_number, columns["查询状态"], "成功")
            worksheet.cell(row_number, columns["错误说明"], "")
            _save_atomic(workbook, self.paths.original)
        finally:
            workbook.close()
        self._refresh_stripped()

    def write_failure(self, record_index: int, record: QueryRecord, reason: str) -> None:
        workbook = load_workbook(self.paths.original)
        try:
            worksheet = workbook["查询结果"]
            columns = _header_map(worksheet)
            row_number = record_index + 2
            worksheet.cell(row_number, columns["查询状态"], "失败")
            worksheet.cell(row_number, columns["错误说明"], str(reason))
            _save_atomic(workbook, self.paths.original)
        finally:
            workbook.close()
        self._refresh_stripped()

    def _refresh_stripped(self) -> None:
        workbook = load_workbook(self.paths.original)
        try:
            worksheet = workbook["查询结果"]
            columns = _header_map(worksheet)
            for header in STRIPPED_TIME_FIELDS:
                column = columns.get(header)
                if not column:
                    continue
                for row_number in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row_number, column)
                    cell.value = strip_minutes(cell.value)
            _save_atomic(workbook, self.paths.stripped)
        finally:
            workbook.close()
