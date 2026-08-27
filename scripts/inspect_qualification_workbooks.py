from pathlib import Path

from openpyxl import load_workbook
from collections import Counter, defaultdict
import re


desktop = Path.home() / "Desktop"
candidates = list(desktop.glob("*.xlsx"))
target_files = [path for path in candidates if "20260814" in path.name]
target_files.extend(
    path for path in candidates
    if 100_000 < path.stat().st_size < 300_000 and path not in target_files
)
qualification_codes = {
    "RAMA", "REUO", "RWAS", "RSEA", "EAMA", "EEUO", "EWAS", "ESEA",
    "RANC", "RORD", "RJFK", "RLAX", "RNLU",
}


def display(value: object) -> str:
    text = "" if value is None else str(value).replace("\r", " ").replace("\n", " ")
    return text[:80]


for workbook_path in target_files:
    print(f"WORKBOOK\t{workbook_path.name}\t{workbook_path.stat().st_size}")
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    for worksheet in workbook.worksheets:
        print(
            f"SHEET\t{worksheet.title}\trows={worksheet.max_row}\tcols={worksheet.max_column}"
            f"\tmerged={len(worksheet.merged_cells.ranges)}\tstate={worksheet.sheet_state}"
        )
        hit_rows: list[int] = []
        for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 120)):
            values = {display(cell.value).strip().upper() for cell in row}
            if len(values & qualification_codes) >= 2:
                hit_rows.append(row[0].row)
        rows_to_show = set(range(1, min(worksheet.max_row, 15) + 1))
        for row_number in hit_rows:
            rows_to_show.update(range(max(1, row_number - 3), min(worksheet.max_row, row_number + 5) + 1))
        for row_number in sorted(rows_to_show):
            cells = []
            for cell in worksheet[row_number]:
                value = display(cell.value)
                if value:
                    cells.append(f"{cell.coordinate}={value}")
            if cells:
                print(f"ROW\t{row_number}\t" + " | ".join(cells[:60]))
        print("TABLES\t" + ",".join(worksheet.tables.keys()))
    print("END")


portal_candidates = [
    path for path in candidates
    if 100_000 < path.stat().st_size < 300_000 and "20260814" not in path.name
]
for workbook_path in portal_candidates:
    print(f"PORTAL_PROFILE\t{workbook_path.name}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        non_empty_rows: list[tuple[int, list[str]]] = []
        distinct_values: dict[int, set[str]] = {}
        for row in worksheet.iter_rows():
            values = [display(cell.value).strip() for cell in row]
            if any(values):
                non_empty_rows.append((row[0].row, values))
            for index, value in enumerate(values):
                if value:
                    distinct_values.setdefault(index, set()).add(value)
        print(f"PORTAL_SHEET\t{worksheet.title}\trows={worksheet.max_row}\tcols={worksheet.max_column}\tnonempty={len(non_empty_rows)}")
        for row_number, values in non_empty_rows[:12]:
            cells = [f"{index + 1}:{value}" for index, value in enumerate(values) if value]
            print(f"PORTAL_ROW\t{row_number}\t" + " | ".join(cells[:40]))
        for index, values in sorted(distinct_values.items()):
            if len(values) <= 25:
                print(f"PORTAL_VALUES\tcol={index + 1}\tcount={len(values)}\t" + " | ".join(sorted(values)))
    print("PORTAL_END")


strength_path = next(path for path in candidates if "20260814" in path.name)
portal_path = next(path for path in candidates if path.name == "运行资质人员名册.xlsx")

strength_book = load_workbook(strength_path, read_only=True, data_only=True)
portal_book = load_workbook(portal_path, read_only=True, data_only=True)

personnel_candidates = []
for sheet in strength_book.worksheets:
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, ())
    normalized = [display(value).strip() for value in header]
    if "员工号" in normalized and "姓名" in normalized:
        code_columns = {
            index: value.upper()
            for index, value in enumerate(normalized)
            if re.fullmatch(r"[RE][A-Z]{3}", value.upper())
        }
        if code_columns:
            personnel_candidates.append((sheet, normalized, code_columns))

print(f"COMPARE_PERSONNEL_CANDIDATES\t{[(sheet.title, len(codes)) for sheet, _, codes in personnel_candidates]}")
personnel_sheet, personnel_header, personnel_code_columns = personnel_candidates[0]
personnel_id_col = personnel_header.index("员工号")
personnel_name_col = personnel_header.index("姓名")
print("COMPARE_PERSONNEL_CODES\t" + " | ".join(personnel_code_columns.values()))

def normalize_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


personnel_by_code: dict[str, dict[str, tuple[str, str, int]]] = defaultdict(dict)
personnel_bad_roles = Counter()
personnel_duplicate_ids = Counter()
personnel_names: dict[str, set[str]] = defaultdict(set)
for row_number, row in enumerate(personnel_sheet.iter_rows(min_row=2, values_only=True), start=2):
    employee_id = normalize_id(row[personnel_id_col] if personnel_id_col < len(row) else None)
    name = display(row[personnel_name_col] if personnel_name_col < len(row) else None).strip()
    if not employee_id and not name:
        continue
    personnel_duplicate_ids[employee_id] += 1
    if name:
        personnel_names[employee_id].add(name)
    for col, code in personnel_code_columns.items():
        value = display(row[col] if col < len(row) else None).strip()
        if not value:
            continue
        if value not in {"1", "2"}:
            personnel_bad_roles[(code, value)] += 1
            continue
        personnel_by_code[code][employee_id] = (name, "机长" if value == "1" else "副驾驶", row_number)

portal_sheet = portal_book[portal_book.sheetnames[0]]
portal_rows = portal_sheet.iter_rows(values_only=True)
portal_header = [display(value).strip() for value in next(portal_rows)]
portal_id_col = portal_header.index("员工号")
portal_name_col = portal_header.index("姓名")
portal_code_col = portal_header.index("资质类别")
portal_by_code: dict[str, dict[str, tuple[str, int]]] = defaultdict(dict)
portal_pair_counts = Counter()
portal_names: dict[str, set[str]] = defaultdict(set)
portal_blank_keys = []
for row_number, row in enumerate(portal_rows, start=2):
    employee_id = normalize_id(row[portal_id_col] if portal_id_col < len(row) else None)
    name = display(row[portal_name_col] if portal_name_col < len(row) else None).strip()
    code = display(row[portal_code_col] if portal_code_col < len(row) else None).strip().upper()
    if not employee_id or not code:
        portal_blank_keys.append((row_number, employee_id, name, code))
        continue
    portal_pair_counts[(code, employee_id)] += 1
    portal_names[employee_id].add(name)
    portal_by_code[code][employee_id] = (name, row_number)

all_codes = sorted(set(personnel_by_code) | set(portal_by_code))
print("COMPARE_PORTAL_CODES\t" + " | ".join(sorted(portal_by_code)))
print(f"COMPARE_BAD_ROLES\t{dict(personnel_bad_roles)}")
print(f"COMPARE_PERSONNEL_DUPLICATE_IDS\t{sum(1 for count in personnel_duplicate_ids.values() if count > 1)}")
print(f"COMPARE_PORTAL_DUPLICATE_PAIRS\t{[(key, count) for key, count in portal_pair_counts.items() if count > 1][:20]}")
print(f"COMPARE_PORTAL_BLANK_KEYS\t{portal_blank_keys[:20]}")
same_id_name_mismatches = []
for employee_id in sorted(set(personnel_names) & set(portal_names)):
    if personnel_names[employee_id] != portal_names[employee_id]:
        same_id_name_mismatches.append((employee_id, sorted(personnel_names[employee_id]), sorted(portal_names[employee_id])))
print(f"COMPARE_NAME_MISMATCHES\t{same_id_name_mismatches[:30]}")
for code in all_codes:
    personnel_ids = set(personnel_by_code[code])
    portal_ids = set(portal_by_code[code])
    only_personnel = sorted(personnel_ids - portal_ids)
    only_portal = sorted(portal_ids - personnel_ids)
    same = personnel_ids & portal_ids
    personnel_sample = [(employee_id, personnel_by_code[code][employee_id][0]) for employee_id in only_personnel[:8]]
    portal_sample = [(employee_id, portal_by_code[code][employee_id][0]) for employee_id in only_portal[:8]]
    print(
        f"COMPARE_CODE\t{code}\tpersonnel={len(personnel_ids)}\tportal={len(portal_ids)}"
        f"\tsame={len(same)}\tonly_personnel={len(only_personnel)}\tonly_portal={len(only_portal)}"
        f"\tpersonnel_sample={personnel_sample}\tportal_sample={portal_sample}"
    )
